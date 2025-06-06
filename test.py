# import canmatrix.formats.ldf
from openpyxl import load_workbook
import pandas as pd
from ldfparser import LDF, LinFrame, LinSignal, LinSignalEncodingType, LinUnconditionalFrame
import canmatrix
import lin
from ctypes import *
from lin.interfaces.peak import PLinApi, LinBus
from lin.interfaces.peak.PLinApi import TLINVersion
from ldfparser import LDF
from ldfparser.frame import LinUnconditionalFrame
from ldfparser.signal import LinSignal
from ldfparser.node import LinMaster, LinSlave
from ldfparser.lin import LinVersion
from ldfparser import save_ldf

# 1. Создаём LDF
ldf = LDF()

# 2. Устанавливаем версию протокола (например LIN 2.1)
ldf._protocol_version = LinVersion(2, 1)
ldf._language_version = LinVersion(2, 1)

# 3. Задаём baudrate и канал
ldf._baudrate = 19200   # Обязательно!
ldf._channel = "LIN1"   # Название канала, можно любое строковое значение

# 4. Создаём мастер и слейв узлы
master = LinMaster(
    name="Master",
    timebase=0.01,
    jitter=0.001,
    max_header_length=8,
    response_tolerance=0.1
)
slave = LinSlave(name="SlaveNode")

# 5. Сохраняем мастер и слейвы в LDF
ldf._master = master
ldf._slaves[slave.name] = slave

# 6. Создаём сигналы
signal1 = LinSignal(name="AmbientLight", width=1, init_value=0)
signal2 = LinSignal(name="InteriorLight", width=1, init_value=0)

# 7. Создаём кадр (Unconditional Frame)
frame = LinUnconditionalFrame(
    frame_id=0x10,
    name="LightControl",
    length=8,
    signals={1: signal1, 2: signal2},
    pad_with_zero=True
)

# 8. Добавляем кадр в LDF
ldf._unconditional_frames[frame.name] = frame

# 9. Сохраняем в файл
save_ldf(ldf, "output.ldf", "C:\\projects\\Convert2DBC\\ldf.jinja2")


# # ldf = canmatrix.CanMatrix()
# protectedId=int("0x97", 16)
# payload=[1, 2, 3, 4, 5, 6]

# ldf = LinBus.LinMessage()

# version = TLINVersion()

# version.Major = 2
# version.Minor = 1
# version.Revision = 5
# version.Build = 1234

# print(version.Major)

# # bus = LinBus()


# ldf.frameId = int("0x17", 16)

# ldf.protectedId = protectedId

# ldf.payload=payload


# print(ldf)


# ldf.protocol = "LIN"

# ldf.add_node("Master") 
# ldf.add_node("Slave1")

# ldf.add_attribute("Master", "ECU1")
# ldf.add_attribute("Slave", "ECU2")
# signal = canmatrix.Signal(
#     name="EngineSpeed",
#     size=8,
#     initial_value=0,
#     min=0,
#     max=255,
#     unit="RPM",
#     receivers=["Slave1"]
# )

# frame = canmatrix.Frame(
#     name="EngineData",
#     arbitration_id=0x10,  
#     size=1,  
#     transmitters=["Master"],
#     receivers=["Slave1"],
# )
# frame.add_signal(signal)

# ldf.add_frame(frame)

# canmatrix.formats.dump(ldf, "example", export_type="ldf")

# def xlsx_to_ldf(input_xlsx: str, output_ldf: str):
#     # Загрузка Excel
#     wb = load_workbook(input_xlsx)
#     df = pd.read_excel(input_xlsx, sheet_name="Matrix", engine="openpyxl", keep_default_na=True)
#     matrix_sheet = wb["Matrix"]
#     df_info = pd.read_excel(input_xlsx, sheet_name="Info", engine="openpyxl", keep_default_na=True)
#     info_sheet = wb["Info"]
#     df_shedule = pd.read_excel(input_xlsx, sheet_name="LIN Schedule", engine="openpyxl", keep_default_na=True)
#     schedule_sheet = wb["LIN Schedule"]

#     baudrate = df_info.iloc[1, 1] * 1000
#     protocol = float(df_info.iloc[1, 0]) if pd.notna(df_info.iloc[1, 0]) else 2.0
#     buses = [col for col in df.columns 
#             if any(val in ["S", "R"] for val in df[col].dropna().unique())]
    
#     ldf = LDF()
#     ldf._protocol_version = protocol
#     ldf._baudrate = baudrate
#     # ldf._master = "DCM_FL"    <-sender   need info
#     # ldf._slaves = ["ALM_FL"]  <-reciever how push Send/Rec
    
#     current_frame = None
#     for num, row in enumerate(matrix_sheet.iter_rows(min_row=2, values_only=True)):
#         if row[0]:
#             publisher="DCM_FL" if row[-2] == "S" else "ALM_FL"
#             length=row[5]
#             current_frame = LinUnconditionalFrame(
#                 frame_id=int(row[1], 16),
#                 name=row[0],
#                 length=row[5],
#                 signals={num: LinSignal("hui", 6, 4)}
#             )
#             ldf.frames.update({row[0]: current_frame})
#     print(current_frame)
#     #         ldf.frames.append(current_frame)
#     #     elif row[6]:
#     #         signal = LinSignal(
#     #             name=row[6],
#     #             start_bit=int(row[9]),
#     #             length=int(row[10]),
#     #             publisher=current_frame.publisher,
#     #             init_value=row[20]
#     #         )
#     #         current_frame.signals.append(signal)

#     # for row in schedule_sheet.iter_rows(min_row=5, values_only=True):
#     #     if row[0]:  # Slot ID
#     #         ldf.add_schedule_table(
#     #             name="LIN_Schedule_1",
#     #             frames=[(row[1], int(row[2]))]
#     #         )

#     # with open(output_ldf, "w") as f:
#     #     f.write(ldf.dump())

# xlsx_to_ldf("ATOM_LIN_Matrix_DCM_FL-ALM_FL_V4.0.0-20250121.xlsx", "output.ldf")