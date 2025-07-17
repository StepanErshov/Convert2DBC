import pandas as pd
from streamlit.runtime.uploaded_file_manager import UploadedFile
from typing import List, Union, Dict
import re
import pprint
import streamlit as st
import os
import math
from openpyxl.worksheet import table
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, 
    Font, 
    Alignment, 
    Border, 
    Side
)
from openpyxl.utils import get_column_letter

st.markdown(
    """
    <style>
    .main {
        background-color: #f5f5f5;
    }
    .stButton>button {
        background-color: #4CAF50;
        color: white;
        border-radius: 5px;
        padding: 10px 24px;
    }
    .stButton>button:hover {
        background-color: #45a049;
    }
    .stFileUploader>div>div>div>button {
        background-color: #2196F3;
        color: white;
    }
    .stTextInput>div>div>input {
        border-radius: 5px;
    }
    .title {
        color: #2c3e50;
    }
    .error-box {
        background-color: #ffebee;
        border-left: 5px solid #f44336;
        padding: 10px;
        margin: 10px 0;
        border-radius: 5px;
    }
    .warning-box {
        background-color: #fff8e1;
        border-left: 5px solid #ffc107;
        padding: 10px;
        margin: 10px 0;
        border-radius: 5px;
    }
    .success-box {
        background-color: #e8f5e9;
        border-left: 5px solid #4caf50;
        padding: 10px;
        margin: 10px 0;
        border-radius: 5px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def get_file_info(file_name: str):
    file_start = "ATOM_CAN_Matrix_"
    file_start1 = "ATOM_CANFD_Matrix_"
    file_name_only = os.path.splitext(os.path.basename(file_name))[0]
    if file_name_only.startswith(file_start1):
        protocol = "CANFD"
        start_index = 0
        parts = file_name_only[len(file_start1) :].split("_")
    elif file_name_only.startswith(file_start):
        protocol = "CAN"
        start_index = 0
        parts = file_name_only[len(file_start) :].split("_")
    else:
        protocol = ""
    if not (
        file_name_only.startswith(file_start) or file_name_only.startswith(file_start1)
    ):
        return None
    start_index = file_name_only.find(file_start1)
    if start_index != -1:
        parts = file_name_only[start_index + len(file_start1) :].split("_")
    else:
        parts = file_name_only[len(file_start) :].split("_")
    domain_name = parts.pop(0)
    version_string = parts.pop(0)
    if version_string.startswith("V"):
        version = version_string[1:]
        versions = version.split(".")
        if len(versions) != 3:
            return None
    else:
        version = ""
    file_date = parts.pop(0)
    if len(parts) > 0:
        if parts[0] == "internal":
            parts.pop(0)
        device_name = "_".join(parts)
    else:
        device_name = ""

    return {
        "version": version,
        "date": file_date,
        "device_name": device_name,
        "domain_name": domain_name,
        "protocol": protocol,
    }


def load_xlsx(file_path: str) -> Union[pd.DataFrame, Dict]:
    try:
        if isinstance(file_path, str) or isinstance(file_path, UploadedFile):
            data_frame = pd.read_excel(
                file_path, sheet_name="Matrix", keep_default_na=True, engine="openpyxl"
            )
            return data_frame
        elif isinstance(file_path, List):
            finally_df = {}
            for file in file_path:
                data_frame = pd.read_excel(
                    file, sheet_name="Matrix", keep_default_na=True, engine="openpyxl"
                )
                if isinstance(file, UploadedFile):
                    finally_df[file.name] = data_frame
                else:
                    finally_df[file.split("\\")[-1]] = data_frame
            return finally_df
    except Exception as e:
        return f"Undefined type of file: {e}"


def create_correct_df(df: pd.DataFrame) -> pd.DataFrame:
    bus_users = [
        col
        for col in df.columns
        if any(val in ["S", "R"] for val in df[col].dropna().unique())
        and col != "Unit\n单位"
    ]
    senders = []
    receivers = []

    for _, row in df.iterrows():
        row_senders = []
        row_receivers = []

        for bus_user in bus_users:
            if bus_user in df.columns:
                if pd.notna(row[bus_user]) and row[bus_user] == "S":
                    row_senders.append(bus_user)
                elif pd.notna(row[bus_user]) and row[bus_user] == "R":
                    row_receivers.append(bus_user)

        senders.append(",".join(row_senders) if row_senders else "Vector__XXX")
        receivers.append(",".join(row_receivers) if row_receivers else "Vector__XXX")

    new_df_data = {
        "Msg ID": df["Msg ID\n报文标识符"].ffill(),
        "Msg Name": df["Msg Name\n报文名称"].ffill(),
        "Cycle Type": df["Msg Cycle Time (ms)\n报文周期时间"].ffill(),
        "Msg Time Fast": df["Msg Cycle Time Fast(ms)\n报文发送的快速周期"].ffill(),
        "Msg Reption": df["Msg Nr. Of Reption\n报文快速发送的次数"].ffill(),
        "Msg Delay": df["Msg Delay Time(ms)\n报文延时时间"].ffill(),
        "Msg Type": df["Msg Type\n报文类型"].ffill(),
        "Send Type": df["Msg Send Type\n报文发送类型"].ffill(),
        "Msg Length": df["Msg Length (Byte)\n报文长度"].ffill(),
        "Sig Name": df["Signal Name\n信号名称"],
        "Start Byte": df["Start Byte\n起始字节"],
        "Start Bit": df["Start Bit\n起始位"],
        "Length": df["Bit Length (Bit)\n信号长度"],
        "Resolution": df["Resolution\n精度"],
        "Offset": df["Offset\n偏移量"],
        "Initinal": df["Initial Value (Hex)\n初始值"],
        "Invalid": df["Invalid Value(Hex)\n无效值"],
        "Min": df["Signal Min. Value (phys)\n物理最小值"],
        "Min Hex": df["Signal Min. Value (Hex)\n总线最小值"],
        "Max": df["Signal Max. Value (phys)\n物理最大值"],
        "Max Hex": df["Signal Max. Value (Hex)\n总线最大值"],
        "Unit": df["Unit\n单位"],
        "Receiver": receivers,
        "Byte Order": df["Byte Order\n排列格式(Intel/Motorola)"],
        "Data Type": df["Data Type\n数据类型"],
        "Description": df["Signal Description\n信号描述"],
        "Signal Value Description": df["Signal Value Description\n信号值描述"],
        "Senders": senders,
        "Signal Send Type": df["Signal Send Type\n信号发送类型"],
        "Inactive value": df["Inactive Value (Hex)\n非使能值"],
    }

    if "BRS\n传输速率切换标识位" in df.columns:
        new_df_data["BRS"] = df["BRS\n传输速率切换标识位"].ffill()
    else:
        new_df_data["BRS"] = None

    if "Frame Format\n帧格式" in df.columns:
        new_df_data["Frame Format"] = df["Frame Format\n帧格式"].ffill()
    else:
        new_df_data["Frame Format"] = None

    new_df = pd.DataFrame(new_df_data)

    new_df["Unit"] = new_df["Unit"].astype(str)
    new_df["Unit"] = new_df["Unit"].str.replace("Ω", "Ohm", regex=False)
    new_df["Unit"] = new_df["Unit"].str.replace("℃", "degC", regex=False)

    new_df = new_df.dropna(subset=["Sig Name"])
    new_df["Is Signed"] = new_df["Data Type"].str.contains("Signed", na=False)

    return new_df


def export_validation_errors_to_excel(data_frame: pd.DataFrame, original_file: Union[str, UploadedFile], output_file_path: str) -> bool:
    all_errors = []

    # 1. Message Name errors
    invalid_names = []
    too_long_names = []
    msg_names = set(data_frame["Msg Name"].dropna().astype(str))
    for name in msg_names:
        if not re.fullmatch(r"^[A-Za-z0-9_\-]+$", name.strip()):
            invalid_names.append(name)
        if len(name) > 64:
            too_long_names.append(name)

    for name in invalid_names:
        all_errors.append(
            {
                "Error Type": "Invalid Message Name",
                "Message/Signal Name": name,
                "Details": "Contains prohibited characters",
                "Expected": "Only A-Z, a-z, 0-9, _, - allowed",
            }
        )

    for name in too_long_names:
        all_errors.append(
            {
                "Error Type": "Too Long Message Name",
                "Message/Signal Name": name,
                "Details": f"Length: {len(name)} characters",
                "Expected": "Max 64 characters",
            }
        )

    # 2. Message Type errors
    msg_type = dict(zip(data_frame["Msg Name"], data_frame["Msg Type"]))
    for name, mtype in msg_type.items():
        if mtype not in ["Normal", "Diag", "NM"]:
            all_errors.append(
                {
                    "Error Type": "Invalid Message Type",
                    "Message/Signal Name": name,
                    "Details": f"Type: {mtype}",
                    "Expected": "Must be Normal, Diag or NM",
                }
            )
        if name.startswith("Diag") and mtype != "Diag":
            all_errors.append(
                {
                    "Error Type": "Message Name-Type Mismatch",
                    "Message/Signal Name": name,
                    "Details": f"Type: {mtype}",
                    "Expected": "Should be Diag for messages starting with 'Diag'",
                }
            )
        if name.startswith("NM_") and mtype != "NM":
            all_errors.append(
                {
                    "Error Type": "Message Name-Type Mismatch",
                    "Message/Signal Name": name,
                    "Details": f"Type: {mtype}",
                    "Expected": "Should be NM for messages starting with 'NM_'",
                }
            )

    # 3. Message ID errors
    data_frame["Msg ID"] = data_frame["Msg ID"].apply(
        lambda x: int(x, 16) if isinstance(x, str) and x.startswith("0x") else int(x)
    )
    msg_id = dict(zip(data_frame["Msg Name"], data_frame["Msg ID"]))
    for name, mid in msg_id.items():
        mtype = msg_type.get(name, "")
        if not (0x001 <= mid <= 0x7FF):
            all_errors.append(
                {
                    "Error Type": "Invalid Message ID",
                    "Message/Signal Name": name,
                    "Details": f"ID: {hex(mid)}",
                    "Expected": "Must be between 0x001 and 0x7FF",
                }
            )
        if 0x700 <= mid <= 0x7FF and mtype != "Diag":
            all_errors.append(
                {
                    "Error Type": "Message ID-Type Mismatch",
                    "Message/Signal Name": name,
                    "Details": f"ID: {hex(mid)}, Type: {mtype}",
                    "Expected": "IDs 0x700-0x7FF should be Diag type",
                }
            )
        if 0x500 <= mid <= 0x5FF and mtype != "NM":
            all_errors.append(
                {
                    "Error Type": "Message ID-Type Mismatch",
                    "Message/Signal Name": name,
                    "Details": f"ID: {hex(mid)}, Type: {mtype}",
                    "Expected": "IDs 0x500-0x5FF should be NM type",
                }
            )

    # 4. Message Send Type errors
    msg_send_type = dict(zip(data_frame["Msg Name"], data_frame["Send Type"]))
    for name, stype in msg_send_type.items():
        if stype not in ["Cycle", "Event", "CE"]:
            all_errors.append(
                {
                    "Error Type": "Invalid Send Type",
                    "Message/Signal Name": name,
                    "Details": f"Send Type: {stype}",
                    "Expected": "Must be Cycle, Event or CE",
                }
            )

    # 5. Frame Format errors (only for CANFD)
    if "Frame Format" in data_frame.columns:
        msg_frame_format = dict(zip(data_frame["Msg Name"], data_frame["Frame Format"]))
        for name, ff in msg_frame_format.items():
            if ff not in ["StandardCAN_FD", "StandardCAN"]:
                all_errors.append(
                    {
                        "Error Type": "Invalid Frame Format",
                        "Message/Signal Name": name,
                        "Details": f"Frame Format: {ff}",
                        "Expected": "Must be StandardCAN_FD or StandardCAN",
                    }
                )

    # 6. BRS errors (only for CANFD)
    if "BRS" in data_frame.columns:
        msg_brs = dict(zip(data_frame["Msg Name"], data_frame["BRS"]))
        msg_frame_format = dict(zip(data_frame["Msg Name"], data_frame["Frame Format"]))
        for name, brs in msg_brs.items():
            ff = msg_frame_format.get(name, "")
            if brs not in [0, 1]:
                all_errors.append(
                    {
                        "Error Type": "Invalid BRS Value",
                        "Message/Signal Name": name,
                        "Details": f"BRS: {brs}",
                        "Expected": "Must be 0 or 1",
                    }
                )
            if brs == 0 and ff != "StandardCAN":
                all_errors.append(
                    {
                        "Error Type": "BRS-Frame Format Mismatch",
                        "Message/Signal Name": name,
                        "Details": f"BRS: {brs}, Frame Format: {ff}",
                        "Expected": "BRS=0 should be with StandardCAN",
                    }
                )
            if brs == 1 and ff != "StandardCAN_FD":
                all_errors.append(
                    {
                        "Error Type": "BRS-Frame Format Mismatch",
                        "Message/Signal Name": name,
                        "Details": f"BRS: {brs}, Frame Format: {ff}",
                        "Expected": "BRS=1 should be with StandardCAN_FD",
                    }
                )

    # 7. Message Length errors
    msg_len = dict(zip(data_frame["Msg Name"], data_frame["Msg Length"]))
    if "Frame Format" in data_frame.columns:
        msg_frame_format = dict(zip(data_frame["Msg Name"], data_frame["Frame Format"]))
        for name, length in msg_len.items():
            ff = msg_frame_format.get(name, "")
            if ff == "StandardCAN_FD" and length not in (8, 64):
                all_errors.append(
                    {
                        "Error Type": "Invalid Message Length",
                        "Message/Signal Name": name,
                        "Details": f"Length: {length}, Frame Format: {ff}",
                        "Expected": "For StandardCAN_FD length must be 8 or 64",
                    }
                )
            elif ff == "StandardCAN" and length != 8:
                all_errors.append(
                    {
                        "Error Type": "Invalid Message Length",
                        "Message/Signal Name": name,
                        "Details": f"Length: {length}, Frame Format: {ff}",
                        "Expected": "For StandardCAN length must be 8",
                    }
                )
    else:
        for name, length in msg_len.items():
            if length != 8:
                all_errors.append(
                    {
                        "Error Type": "Invalid Message Length",
                        "Message/Signal Name": name,
                        "Details": f"Length: {length}",
                        "Expected": "For CAN length must be 8",
                    }
                )

    # 8. Signal Name errors
    invalid_sig_names = []
    too_long_sig_names = []
    need_change_sig_names = []
    sig_name = set(data_frame["Sig Name"].dropna().astype(str))
    for name in sig_name:
        if not re.fullmatch(r"^[A-Za-z0-9_\-]+$", name.strip()):
            invalid_sig_names.append(name)
        if len(name) > 64:
            too_long_sig_names.append(name)
        if len(name) > 36 and len(name) < 64:
            need_change_sig_names.append(name)

    for name in invalid_sig_names:
        all_errors.append(
            {
                "Error Type": "Invalid Signal Name",
                "Message/Signal Name": name,
                "Details": "Contains prohibited characters",
                "Expected": "Only A-Z, a-z, 0-9, _, - allowed",
            }
        )

    for name in too_long_sig_names:
        all_errors.append(
            {
                "Error Type": "Too Long Signal Name",
                "Message/Signal Name": name,
                "Details": f"Length: {len(name)} characters",
                "Expected": "Max 64 characters",
            }
        )

    for name in need_change_sig_names:
        all_errors.append(
            {
                "Error Type": "Signal Name Needs Shortening",
                "Message/Signal Name": name,
                "Details": f"Length: {len(name)} characters",
                "Expected": "Recommended max 36 characters",
            }
        )

    # 9. Signal Value Description errors
    sig_desc = dict(zip(data_frame["Sig Name"], data_frame["Signal Value Description"]))
    pattern = re.compile(
        r"^(?:"
        r"0x[0-9A-Fa-f]+(:|~0x[0-9A-Fa-f]+:)\s*"
        r"[<>A-Za-z0-9 _+\-.,/%°()&;]+"
        r"(?:\s*[+&]\s*[<>A-Za-z0-9 _+\-.,/%°()]+)*"
        r"(?:\n|$)"
        r")+$"
    )
    for sig_name, val in sig_desc.items():
        if pd.isna(val):
            all_errors.append(
                {
                    "Error Type": "Missing Signal Value Description",
                    "Message/Signal Name": sig_name,
                    "Details": "Value is empty",
                    "Expected": "Signal value description is required",
                }
            )
            continue
        str_val = str(val).strip()
        if not pattern.fullmatch(str_val):
            all_errors.append(
                {
                    "Error Type": "Invalid Signal Value Description",
                    "Message/Signal Name": sig_name,
                    "Details": f"Value: {str_val}",
                    "Expected": "Must match pattern like '0x0: No Error' or '0x0~0x3: Reserved'",
                }
            )

    # 10. Signal Description errors
    sig_desc = dict(zip(data_frame["Sig Name"], data_frame["Description"]))
    for sig_name, val in sig_desc.items():
        if pd.isna(val):
            # all_errors.append(
            #     {
            #         "Error Type": "Missing Signal Description",
            #         "Message/Signal Name": sig_name,
            #         "Details": "Value is empty",
            #         "Expected": "Signal description is required",
            #     }
            # )
            continue
        str_val = str(val)
        if not re.fullmatch(r"^[A-Za-z0-9 ,.;:+_/-<>%()°~-]+$", str_val):
            all_errors.append(
                {
                    "Error Type": "Invalid Signal Description",
                    "Message/Signal Name": sig_name,
                    "Details": f"Value: {str_val}",
                    "Expected": "Contains invalid characters",
                }
            )

    # 11. Byte Order errors
    byte_order = dict(zip(data_frame["Sig Name"], data_frame["Byte Order"]))
    for sig_name, byte in byte_order.items():
        if byte != "Motorola MSB":
            all_errors.append(
                {
                    "Error Type": "Invalid Byte Order",
                    "Message/Signal Name": sig_name,
                    "Details": f"Byte Order: {byte}",
                    "Expected": "Must be 'Motorola MSB'",
                }
            )

    # 12. Start Byte errors
    start_byte = dict(zip(data_frame["Sig Name"], data_frame["Start Byte"]))
    for sig_name, byte in start_byte.items():
        if byte not in range(0, 8):
            all_errors.append(
                {
                    "Error Type": "Invalid Start Byte",
                    "Message/Signal Name": sig_name,
                    "Details": f"Start Byte: {byte}",
                    "Expected": "Must be between 0 and 7",
                }
            )

    # 13. Start Bit errors
    start_bit = dict(zip(data_frame["Sig Name"], data_frame["Start Bit"]))
    for sig_name, bit in start_bit.items():
        if bit not in range(0, 64):
            all_errors.append(
                {
                    "Error Type": "Invalid Start Bit",
                    "Message/Signal Name": sig_name,
                    "Details": f"Start Bit: {bit}",
                    "Expected": "Must be between 0 and 63",
                }
            )

    # 14. Signal Send Type errors
    sig_send_type = dict(zip(data_frame["Sig Name"], data_frame["Signal Send Type"]))
    msg_send_type = dict(zip(data_frame["Msg Name"], data_frame["Send Type"]))
    validation_rules = {
        "CA": ["Cycle", "IfActiveWithRepetition"],
        "CE": [
            "Cycle",
            "OnWrite",
            "OnChange",
            "OnWriteWithRepetition",
            "OnChangeWithRepetition",
        ],
        "Cycle": ["Cycle"],
        "Event": [
            "OnWrite",
            "OnChange",
            "OnWriteWithRepetition",
            "OnChangeWithRepetition",
        ],
        "IfActive": ["IfActive"],
    }
    for sig_name, sig_type in sig_send_type.items():
        msg_name = data_frame[data_frame["Sig Name"] == sig_name]["Msg Name"].iloc[0]
        msg_type = msg_send_type.get(msg_name)
        if msg_type in validation_rules:
            allowed_types = validation_rules[msg_type]
            if sig_type not in allowed_types:
                all_errors.append(
                    {
                        "Error Type": "Invalid Signal Send Type",
                        "Message/Signal Name": sig_name,
                        "Details": f"Signal Type: {sig_type}, Message Type: {msg_type}",
                        "Expected": f"Allowed types: {', '.join(allowed_types)}",
                    }
                )

    # 15. Resolution errors
    resol = dict(zip(data_frame["Sig Name"], data_frame["Resolution"]))
    for sig_name, res in resol.items():
        if pd.isna(res):
            all_errors.append(
                {
                    "Error Type": "Missing Resolution",
                    "Message/Signal Name": sig_name,
                    "Details": "Value is empty",
                    "Expected": "Resolution is required",
                }
            )
            continue
        if type(res) not in [int, float]:
            all_errors.append(
                {
                    "Error Type": "Invalid Resolution Type",
                    "Message/Signal Name": sig_name,
                    "Details": f"Type: {type(res).__name__}, Value: {res}",
                    "Expected": "Must be int or float",
                }
            )

    # 16. Offset errors
    offset = dict(zip(data_frame["Sig Name"], data_frame["Offset"]))
    for sig_name, off in offset.items():
        if pd.isna(off):
            all_errors.append(
                {
                    "Error Type": "Missing Offset",
                    "Message/Signal Name": sig_name,
                    "Details": "Value is empty",
                    "Expected": "Offset is required",
                }
            )
            continue
        if type(off) not in [int, float]:
            all_errors.append(
                {
                    "Error Type": "Invalid Offset Type",
                    "Message/Signal Name": sig_name,
                    "Details": f"Type: {type(off).__name__}, Value: {off}",
                    "Expected": "Must be int or float",
                }
            )

    # 17. Minimum value errors
    min_phys = dict(zip(data_frame["Sig Name"], data_frame["Min"]))
    min_hex = dict(zip(data_frame["Sig Name"], data_frame["Min Hex"]))
    resolutions = dict(zip(data_frame["Sig Name"], data_frame["Resolution"]))
    offsets = dict(zip(data_frame["Sig Name"], data_frame["Offset"]))
    for sig_name in min_phys.keys():
        phys = min_phys.get(sig_name)
        hex_val = min_hex.get(sig_name)
        res = resolutions.get(sig_name)
        offset_val = offsets.get(sig_name, 0)

        if pd.isna(phys) or pd.isna(hex_val) or pd.isna(res):
            continue

        try:
            if isinstance(hex_val, str):
                if hex_val.startswith("0x"):
                    hex_int = int(hex_val, 16)
                else:
                    hex_int = int(hex_val)
            else:
                hex_int = int(hex_val)

            calculated_phys = hex_int * res + offset_val

            if not math.isclose(calculated_phys, phys, rel_tol=1e-9):
                all_errors.append(
                    {
                        "Error Type": "Invalid Minimum Value Calculation",
                        "Message/Signal Name": sig_name,
                        "Details": f"Min (Physical): {phys}, Min (Hex): {hex_val}, Calculated: {calculated_phys}",
                        "Expected": "Physical should equal (Hex * Resolution) + Offset",
                    }
                )
        except (ValueError, TypeError) as e:
            all_errors.append(
                {
                    "Error Type": "Invalid Minimum Value Format",
                    "Message/Signal Name": sig_name,
                    "Details": f"Min (Hex): {hex_val}, Error: {str(e)}",
                    "Expected": "Hex value should be convertible to integer",
                }
            )

    # 18. Maximum value errors
    max_phys = dict(zip(data_frame["Sig Name"], data_frame["Max"]))
    max_hex = (
        dict(zip(data_frame["Sig Name"], data_frame["Max Hex"]))
        if "Max Hex" in data_frame.columns
        else {}
    )
    if not max_hex:
        max_hex = dict(zip(data_frame["Sig Name"], data_frame["Invalid"]))

    invalid_sig_val_desc = {}
    last_sign_val_desc = dict(zip(data_frame["Sig Name"], data_frame["Signal Value Description"]))

    for sig_name, val in last_sign_val_desc.items():
        if pd.isna(val):
            invalid_sig_val_desc[sig_name] = "NaN"
            continue
        
        str_val = str(val).strip().split("\n")[-1].split(":")[0].split("~")[-1]
        try:
            num = int(str_val, 16)
            if max_phys[sig_name] < num:
                all_errors.append ({
                    "Error Type": "Phys/Hex max value greater than max signal value description",
                    "Message/Signal Name": sig_name,
                    "Details": f"Max (Physical): {max_phys[sig_name]}, Max (Hex): {max_hex[sig_name]}, Signal Value Description: {num}",
                    "Expected": "The maximum Phys/Hex value must be greater than or equal to the last value in the signal value description."
                })
        except ValueError:
            continue

    for sig_name in max_phys.keys():
        phys = max_phys.get(sig_name)
        hex_val = max_hex.get(sig_name)
        res = resolutions.get(sig_name)
        offset_val = offsets.get(sig_name, 0)

        if pd.isna(phys) or pd.isna(hex_val) or pd.isna(res):
            continue

        try:
            if isinstance(hex_val, str):
                if hex_val.startswith("0x"):
                    hex_int = int(hex_val, 16)
                else:
                    hex_int = int(hex_val)
            else:
                hex_int = int(hex_val)

            calculated_phys = hex_int * res + offset_val

            if not math.isclose(calculated_phys, phys, rel_tol=1e-9):
                all_errors.append(
                    {
                        "Error Type": "Invalid Maximum Value Calculation",
                        "Message/Signal Name": sig_name,
                        "Details": f"Max (Physical): {phys}, Max (Hex): {hex_val}, Calculated: {calculated_phys}",
                        "Expected": "Physical should equal (Hex * Resolution) + Offset",
                    }
                )
        except (ValueError, TypeError) as e:
            all_errors.append(
                {
                    "Error Type": "Invalid Maximum Value Format",
                    "Message/Signal Name": sig_name,
                    "Details": f"Max (Hex): {hex_val}, Error: {str(e)}",
                    "Expected": "Hex value should be convertible to integer",
                }
            )
    

    # Initinal
    signal_lengths = dict(zip(data_frame["Sig Name"], data_frame["Length"].astype(int)))
    max_values = dict(zip(data_frame["Sig Name"], data_frame["Max"]))
    
    try:
        initial_values = dict(zip(
            data_frame["Sig Name"], 
            data_frame["Initinal"].apply(
                lambda x: int(x, 16) if isinstance(x, str) and x.startswith(("0x", "0X")) 
                else int(x) if isinstance(x, str) 
                else x
            )
        ))
        
        invalid_values = dict(zip(
            data_frame["Sig Name"], 
            data_frame["Invalid"].apply(
                lambda x: int(x, 16) if isinstance(x, str) and x.startswith(("0x", "0X")) 
                else int(x) if isinstance(x, str) 
                else x
            )
        ))
    except (ValueError, AttributeError) as e:
        st.error(f"Failed to parse Initial or Invalid values: {str(e)}")
        return False

    invalid_signals = []

    for sig_name, length in signal_lengths.items():
        if pd.isna(length):
            continue
            
        max_allowed = (1 << int(length)) - 1
        
        max_val = max_values.get(sig_name)
        init_val = initial_values.get(sig_name)
        inval_val = invalid_values.get(sig_name)

        if not pd.isna(max_val) and max_val > max_allowed:
            all_errors.append({
                "Error Type": "Max value exceeding bit length limits",
                "Message/Signal Name": sig_name,
                "Details": f"Max Value: {max_val}, Bit Length: {length}, Max Allowed: {max_allowed}",
                "Expected": "Signal values (Max, Initial, Invalid) must not exceed 2^N - 1, where N is the signal bit length",
            })

        if not pd.isna(init_val) and init_val > max_allowed:
            all_errors.append({
                "Error Type": "Initinal value exceeding bit length limits",
                "Message/Signal Name": sig_name,
                "Details": f"Initinal Value: {init_val}, Bit Length: {length}, Max Allowed: {max_allowed}",
                "Expected": "Signal values (Max, Initial, Invalid) must not exceed 2^N - 1, where N is the signal bit length",
            })

        if not pd.isna(inval_val) and inval_val > max_allowed:
            invalid_signals.append({
                "Error Type": "Invalid value exceeding bit length limits",
                "Message/Signal Name": sig_name,
                "Details": f"Invalid Value: {inval_val}, Bit Length: {length}, Max Allowed: {max_allowed}",
                "Expected": "Signal values (Max, Initial, Invalid) must not exceed 2^N - 1, where N is the signal bit length",
            })


    if not all_errors:
        return False

    if isinstance(original_file, UploadedFile):
        temp_path = "temp_input.xlsx"
        with open(temp_path, "wb") as f:
            f.write(original_file.getbuffer())
        wb = load_workbook(temp_path)
        os.remove(temp_path)
    else:
        wb = load_workbook(original_file)
    
    all_sheet_name = ["Cover", "History", "Data ID", "Legend", "CheckResult", "ChangeList"]

    for sheet in all_sheet_name:
        del wb[sheet]

    ws = wb["Matrix"]

    error_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    warning_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bg_fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")
    error_font = Font(color="FFFFFF", bold=True)
    
    header_map = {}
    for cell in ws[1]:
        header_map[cell.value] = cell.column

    column_mapping = {
        "Msg Name": "Msg Name\n报文名称",
        "Msg Type": "Msg Type\n报文类型",
        "Msg ID": "Msg ID\n报文标识符",
        "Cycle Type": "Msg Cycle Time (ms)\n报文周期时间",
        "Msg Time Fast": "Msg Cycle Time Fast(ms)\n报文发送的快速周期",
        "Msg Reption": "Msg Nr. Of Reption\n报文快速发送的次数",
        "Msg Delay": "Msg Delay Time(ms)\n报文延时时间",
        "Send Type": "Msg Send Type\n报文发送类型",
        "Msg Length": "Msg Length (Byte)\n报文长度",
        "Sig Name": "Signal Name\n信号名称",
        "Start Byte": "Start Byte\n起始字节",
        "Start Bit": "Start Bit\n起始位",
        "Length": "Bit Length (Bit)\n信号长度",
        "Resolution": "Resolution\n精度",
        "Offset": "Offset\n偏移量",
        "Initinal": "Initial Value (Hex)\n初始值",
        "Invalid": "Invalid Value(Hex)\n无效值",
        "Min": "Signal Min. Value (phys)\n物理最小值",
        "Min Hex": "Signal Min. Value (Hex)\n总线最小值",
        "Max": "Signal Max. Value (phys)\n物理最大值",
        "Max Hex": "Signal Max. Value (Hex)\n总线最大值",
        "Unit": "Unit\n单位",
        "Byte Order": "Byte Order\n排列格式(Intel/Motorola)",
        "Data Type": "Data Type\n数据类型",
        "Description": "Signal Description\n信号描述",
        "Signal Value Description": "Signal Value Description\n信号值描述",
        "Signal Send Type": "Signal Send Type\n信号发送类型",
        "Inactive value": "Inactive Value (Hex)\n非使能值",
        "Frame Format": "Frame Format\n帧格式",
        "BRS": "BRS\n传输速率切换标识位"
    }
    
    row_map = {}
    for row_idx in range(2, ws.max_row + 1):
        msg_name = ws.cell(row=row_idx, column=header_map[column_mapping["Msg Name"]]).value
        sig_name = ws.cell(row=row_idx, column=header_map[column_mapping["Sig Name"]]).value
        if msg_name or sig_name:
            row_map[(str(msg_name).strip(), str(sig_name).strip())] = row_idx


    for error in all_errors:
        error_type = error["Error Type"]
        name = error["Message/Signal Name"].strip()

        column_key = None

        if "Message Name" in error_type:
            column_key = "Msg Name"
        elif "Message Type" in error_type:
            column_key = "Msg Type"
        elif "Message ID" in error_type:
            column_key = "Msg ID"
        elif "Send Type" in error_type:
            column_key = "Send Type"
        elif "Frame Format" in error_type:
            column_key = "Frame Format"
        elif "BRS" in error_type:
            column_key = "BRS"
        elif "Message Length" in error_type:
            column_key = "Msg Length"
        elif "Signal Name" in error_type:
            column_key = "Sig Name"
        elif "Signal Value Description" in error_type:
            column_key = "Signal Value Description"
        elif "Signal Description" in error_type:
            column_key = "Description"
        elif "Byte Order" in error_type:
            column_key = "Byte Order"
        elif "Start Byte" in error_type:
            column_key = "Start Byte"
        elif "Start Bit" in error_type:
            column_key = "Start Bit"
        elif "Signal Send Type" in error_type:
            column_key = "Signal Send Type"
        elif "Resolution" in error_type:
            column_key = "Resolution"
        elif "Offset" in error_type:
            column_key = "Offset"
        elif "Minimum" in error_type:
            column_key = "Min"
        elif "Maximum" in error_type:
            column_key = "Max"
        
        if not column_key or column_key not in column_mapping:
            continue
        
        for (msg_name, sig_name), row_idx in row_map.items():
            if msg_name == name or sig_name == name:
                col_name = column_mapping[column_key]
                if col_name in header_map:
                    col_idx = header_map[col_name]
                    ws.cell(row=row_idx, column=col_idx).fill = error_fill
                    ws.cell(row=row_idx, column=col_idx).font = error_font

    
    # ws = wb["CheckResult"]

    if "CheckResult" not in wb.sheetnames:
        wb.create_sheet("CheckResult")
    
    ws = wb["CheckResult"]
    
    header_style = {
    'fill': PatternFill(start_color="00CCFF", fill_type="solid"),
    'font': Font(name='宋体', bold=True, size=13, color='000000', italic=True),
    'alignment': Alignment(horizontal="center", vertical="center"),
    'border': Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
        )
    }

    column_names = ["Serial number", "Warning site", "Warning description", "Expected"]
    for col_num, name in enumerate(column_names, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = name
        for attr, value in header_style.items():
            setattr(cell, attr, value)
    
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = len(name) + 20
    
    start_col = 1
    end_col = len(column_names)
    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)

    merged_cell = ws.cell(row=2, column=start_col)
    merged_cell.value = "【Warning】All warning cells"
    for attr, value in header_style.items():
        setattr(merged_cell, attr, value)

    for row_idx, error in enumerate(all_errors, start=3):
        cell1 = ws.cell(row=row_idx, column=1)
        cell1.value = row_idx - 2
        cell1.font = Font(name='宋体', size=12, color='000000')
        cell1.alignment = Alignment(horizontal="center", vertical="center")
        cell2 = ws.cell(row=row_idx, column=2)
        cell2.value = error["Message/Signal Name"]
        cell2.font = Font(name='宋体', size=12, color='000000')
        cell2.alignment = Alignment(horizontal="center", vertical="center")
        cell3 = ws.cell(row=row_idx, column=3)
        cell3.value = error["Details"]
        cell3.font = Font(name='宋体', size=12, color='000000')
        cell4 = ws.cell(row=row_idx, column=4)
        cell4.value = error["Expected"]
        cell4.font = Font(name="宋体", size=12, color="000000")
            
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(column_names)):
        for cell in row:
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    wb.save(output_file_path)

    return True


def get_column_index(ws, column_name):
    for cell in ws[1]:
        if cell.value == column_name:
            return cell.column
    return None


def validate_messages_name(data_frame: pd.DataFrame) -> bool:
    invalid_names = []
    too_long_names = []

    msg_names = set(data_frame["Msg Name"].dropna().astype(str))

    for name in msg_names:
        if not re.fullmatch(r"^[A-Za-z0-9_\-]+$", name.strip()):
            invalid_names.append(name)

        if len(name) > 64:
            too_long_names.append(name)

    if not invalid_names and not too_long_names:
        st.success("All message titles are correct!")
        return True

    if invalid_names:
        with st.expander(
            "Incorrect names (contain prohibited characters)", expanded=True
        ):
            st.error(f"Found {len(invalid_names)} incorrect name:")
            st.dataframe(pd.DataFrame({"Incorrect name": invalid_names}))
            st.info("Allowed characters: A-Z, a-z, 0-9, _, -")

    if too_long_names:
        with st.expander("Names too long (>64 characters)", expanded=True):
            st.warning(f"Found {len(too_long_names)} too long name:")
            st.dataframe(
                pd.DataFrame(
                    {"Name": too_long_names, "Len": [len(n) for n in too_long_names]}
                )
            )

    return False


def validate_messages_type(data_frame: pd.DataFrame) -> bool:
    msg_type = dict(zip(data_frame["Msg Name"], data_frame["Msg Type"]))

    invalid_type = {}
    invalid_name = {}

    for key, val in msg_type.items():
        if val not in ["Normal", "Diag", "NM"]:
            invalid_type[key] = val

        if key.startswith("Diag") and val != "Diag":
            invalid_name[key] = val

        if key.startswith("NM_") and val != "NM":
            invalid_name[key] = val

    if not invalid_name and not invalid_type:
        st.success("All message types are correct!")
        return True

    if invalid_type:
        with st.expander("Incorrect type (Unknown type)", expanded=True):
            st.error(f"Found {len(invalid_type.keys())} incorrect type:")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Mes Name": invalid_type.keys(),
                        "Incorrect types": invalid_type.values(),
                    }
                )
            )
            st.info("list of allowed values ​​'Normal', 'Diag', 'NM'")

    if invalid_name:
        with st.expander("Incorrect name (Not for this type))", expanded=True):
            st.error(f"Found {len(invalid_name.keys())} incorrect type:")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Incorrect Name": invalid_name.keys(),
                        "Msg Type": invalid_name.values(),
                    }
                )
            )
            st.info(
                "NM, if Msg Name first 3 characters = 'NM_' and Diag, if Msg Name firsts 4 characters = 'Diag'"
            )

    return False


def validate_messages_id(data_frame: pd.DataFrame) -> bool:
    data_frame["Msg ID"] = data_frame["Msg ID"].apply(
        lambda x: int(x, 16) if isinstance(x, str) and x.startswith("0x") else int(x)
    )

    msg_id = dict(zip(data_frame["Msg Name"], data_frame["Msg ID"]))
    msg_type = dict(zip(data_frame["Msg Name"], data_frame["Msg Type"]))

    invalid_id = {}
    invalid_type = {}

    for mes, id in msg_id.items():
        if not (0x001 <= id <= 0x7FF):
            invalid_id[mes] = id
        if 0x700 <= id <= 0x7FF and msg_type[mes] != "Diag":
            invalid_type[mes] = id
        if 0x500 <= id <= 0x5FF and msg_type[mes] != "NM":
            invalid_type[mes] = id

    if not invalid_type and not invalid_id:
        st.success("All message IDs are correct!")
        return True

    if invalid_id:
        with st.expander(
            "Incorrect ID (Whether it fits within the range or not))", expanded=True
        ):
            st.error(f"Found {len(invalid_id.keys())} incorrect IDs:")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Msg Name": invalid_id.keys(),
                        "Incorrect IDs": invalid_id.values(),
                    }
                )
            )
            st.info("Msg ID - Must be in the range 0x001 to 0x7FF (Hex)")

    if invalid_type:
        with st.expander("Incorrect ID for Msg Type (Wrong range)", expanded=True):
            st.error(f"Found {len(invalid_type.keys())} incorrect types:")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Msg Name": invalid_type.keys(),
                        "Incorrect IDs": invalid_type.values(),
                    }
                )
            )
            st.info(
                "Diag if Message ID is in the range 0x700 to 7FF and NM if Message ID is in the range 0x500 to 5FF"
            )

    return False


def validate_messages_send_type(data_frame: pd.DataFrame) -> bool:

    msg_send_type = dict(zip(data_frame["Msg Name"], data_frame["Send Type"]))

    invalid_send_type = {}

    for mes, send_type in msg_send_type.items():
        if send_type not in ["Cycle", "Event", "CE"]:
            invalid_send_type[mes] = send_type

    if not invalid_send_type:
        st.success("All messages send types are correct!")
        return True

    if invalid_send_type:
        with st.expander("Incorrect send types()", expanded=True):
            st.error(f"Found {len(invalid_send_type.keys())} incorrect send types:")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Msg Name": invalid_send_type.keys(),
                        "Incorrect type": invalid_send_type.values(),
                    }
                )
            )
            st.info("Send Type should be 'Cycle', 'Event' or 'CE'")

    return False


def validate_messages_frame_fromat(
    file_path: Union[UploadedFile, str, List], data_frame: pd.DataFrame
) -> bool:
    file_info = get_file_info(file_path)

    if file_info["protocol"] != "CANFD":
        st.warning(
            f"Frame Format validation is not applicable for {file_info['protocol']} protocol"
        )
        return True

    if "Frame Format" not in data_frame.columns:
        st.error("Frame Format column not found in the dataframe")
        return False
    msg_frame_format = dict(zip(data_frame["Msg Name"], data_frame["Frame Format"]))

    invalid_ff = {}

    for mes, ff in msg_frame_format.items():
        if ff not in ["StandardCAN_FD", "StandardCAN"]:
            invalid_ff[mes] = ff

    if not invalid_ff:
        st.success("All messages frame formats are correct!")
        return True

    if invalid_ff:
        st.error(f"Found {len(invalid_ff.keys())} incorrect frame formats:")
        st.dataframe(
            pd.DataFrame(
                {
                    "Msg Name": invalid_ff.keys(),
                    "Incorrect frame format": invalid_ff.values(),
                }
            )
        )
        st.info("Frame format should be 'StandardCAN_FD' or 'StandardCAN'")

    return False


def validate_messages_BRS(
    file_path: Union[UploadedFile, str, List], data_frame: pd.DataFrame
) -> bool:
    file_info = get_file_info(file_path)

    if file_info["protocol"] != "CANFD":
        st.warning(
            f"BRS validation is not applicable for {file_info['protocol']} protocol"
        )
        return True

    if "BRS" not in data_frame.columns:
        st.error("BRS column not found in the dataframe")
        return False

    msg_brs = dict(zip(data_frame["Msg Name"], data_frame["BRS"]))
    msg_frame_format = dict(zip(data_frame["Msg Name"], data_frame["Frame Format"]))

    invalid_brs = {}
    invalid_brs_protocol = {}

    for mes, brs in msg_brs.items():
        if brs not in [0, 1]:
            invalid_brs[mes] = brs

        if brs == 0 and msg_frame_format[mes] != "StandardCAN":
            invalid_brs_protocol[mes] = {
                "BRS": brs,
                "Frame Format": msg_frame_format[mes],
            }

        if brs == 1 and msg_frame_format[mes] != "StandardCAN_FD":
            invalid_brs_protocol[mes] = {
                "BRS": brs,
                "Frame Format": msg_frame_format[mes],
            }

    if not invalid_brs and not invalid_brs_protocol:
        st.success("All BRS values are correct!")
        return True

    if invalid_brs:
        with st.expander("Incorrect BRS values", expanded=True):
            st.error(f"Found {len(invalid_brs)} incorrect BRS values:")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Msg Name": invalid_brs.keys(),
                        "Incorrect BRS": invalid_brs.values(),
                    }
                )
            )
            st.info("BRS should be '1' or '0'")

    if invalid_brs_protocol:
        with st.expander("Incorrect BRS for Frame Format", expanded=True):
            st.error(
                f"Found {len(invalid_brs_protocol)} incorrect BRS for Frame Format"
            )
            df_data = []
            for msg_name, values in invalid_brs_protocol.items():
                df_data.append(
                    {
                        "Msg Name": msg_name,
                        "Incorrect BRS": values["BRS"],
                        "Frame Format": values["Frame Format"],
                    }
                )
            st.dataframe(pd.DataFrame(df_data))
            st.info(
                "BRS=0 should be with StandardCAN, BRS=1 should be with StandardCAN_FD"
            )

    return False


def validate_messages_length(
    file: Union[str, UploadedFile], data_frame: pd.DataFrame
) -> bool:
    file_info = get_file_info(file_name=file)
    protocol = file_info.get("protocol", "").upper()

    msg_len = dict(zip(data_frame["Msg Name"], data_frame["Msg Length"]))
    invalid_len = {}

    if "CANFD" in protocol:
        msg_frame_format = dict(zip(data_frame["Msg Name"], data_frame["Frame Format"]))

        for mes, length in msg_len.items():
            frame_format = msg_frame_format[mes]

            if frame_format == "StandardCAN_FD":
                if length not in (8, 64):
                    invalid_len[mes] = {"Len": length, "Frame": frame_format}
            elif frame_format == "StandardCAN":
                if length != 8:
                    invalid_len[mes] = {"Len": length, "Frame": frame_format}
            else:
                invalid_len[mes] = {"Len": length, "Frame": frame_format}

    else:
        for mes, length in msg_len.items():
            if length != 8:
                invalid_len[mes] = {"Len": length}

    if not invalid_len:
        st.success("All messages length are correct!")
        return True

    if invalid_len:
        with st.expander("Incorrect messages length", expanded=True):
            st.error(f"Found {len(invalid_len.keys())} incorrect length")
            df_data = []

            for msg_name, values in invalid_len.items():
                record = {"Msg Name": msg_name, "Incorrect Length": values["Len"]}
                if "CANFD" in protocol:
                    record["Frame Format"] = values.get("Frame", "N/A")
                df_data.append(record)

            st.dataframe(pd.DataFrame(df_data))

            if "CANFD" in protocol:
                st.info(
                    """
                For CAN FD messages:
                - StandardCAN_FD: length must be 8 or 64 bytes
                - StandardCAN: length must be 8 bytes
                """
                )
            else:
                st.info("For CAN, message length must be 8 bytes.")

    return False


def validate_signal_names(data_frame: pd.DataFrame) -> bool:
    invalid_names = []
    too_long_names = []
    need_change = []
    sig_name = set(data_frame["Sig Name"].dropna().astype(str))

    for name in sig_name:
        if not re.fullmatch(r"^[A-Za-z0-9_\-]+$", name.strip()):
            invalid_names.append(name)

        if len(name) > 64:
            too_long_names.append(name)
        if len(name) > 36 and len(name) < 64:
            need_change.append(name)

    if not invalid_names and not too_long_names and not need_change:
        st.success("All signals titles are correct!")
        return True

    if invalid_names:
        with st.expander(
            "Incorrect names (contain prohibited characters)", expanded=True
        ):
            st.error(f"Found {len(invalid_names)} incorrect name:")
            st.dataframe(pd.DataFrame({"Incorrect name": invalid_names}))
            st.info("Allowed characters: A-Z, a-z, 0-9, _, -")

    if too_long_names:
        with st.expander("Names too long (>64 characters)", expanded=True):
            st.warning(f"Found {len(too_long_names)} too long name:")
            st.dataframe(
                pd.DataFrame(
                    {"Name": too_long_names, "Len": [len(n) for n in too_long_names]}
                )
            )

    if need_change:
        with st.expander("Names, which need change (>36 characters)", expanded=True):
            st.warning(f"Found {len(need_change)} need change name:")
            st.dataframe(
                pd.DataFrame(
                    {"Name": need_change, "Len": [len(n) for n in need_change]}
                )
            )
            st.info("Please, try to make the Signal name shorter")

    return False


def validate_signal_value_description(data_frame: pd.DataFrame) -> bool:
    sig_desc = dict(zip(data_frame["Sig Name"], data_frame["Signal Value Description"]))

    invalid_val = {}
    invalid_nan = {}
    invalid_chars = {}

    pattern = re.compile(
        r"^(?:"
        r"0x[0-9A-Fa-f]+(:|~0x[0-9A-Fa-f]+:)\s*"
        r"[<>A-Za-z0-9 _+\-.,/%°()&]+"
        r"(?:\s*[+&]\s*[<>A-Za-z0-9 _+\-.,/%°()]+)*"
        r"(?:\n|$)"
        r")+$"
    )

    allowed_chars_pattern = re.compile(r'^[A-Za-z0-9 ,.:+_/\-<>%()&~-]+$')
    for sig_name, val in sig_desc.items():
        if pd.isna(val):
            invalid_nan[sig_name] = "NaN"
            continue

        str_val = str(val).strip()

        parts = re.split(r'(0x[0-9A-Fa-f]+[:~]?)', str_val)
        descriptions = [part for i, part in enumerate(parts) if i % 2 == 0 and part.strip()]
        
        has_invalid_chars = False
        for desc in descriptions:
            if not allowed_chars_pattern.match(desc):
                has_invalid_chars = True
                break
                
        if has_invalid_chars:
            invalid_chars[sig_name] = str_val
            continue
            
        if not pattern.fullmatch(str_val):
            invalid_val[sig_name] = str_val

    if not invalid_nan and not invalid_val and not invalid_chars:
        st.success("All Signal Values Description are correct!")
        return True

    if invalid_chars:
        with st.expander(
            "Invalid characters in signal value description", expanded=True
        ):
            st.error(f"Found {len(invalid_chars)} value descriptions with invalid characters")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Signal Name": invalid_chars.keys(),
                        "Invalid Description": invalid_chars.values(),
                    }
                )
            )
            st.info(
                "Allowed characters are: A-Z, a-z, 0-9, spaces, and the following symbols: ,.:+_/-<>%()~&°\n"
                "Examples of invalid characters: !@#$^=[]{}|'\" etc."
            )

    if invalid_val:
        with st.expander(
            "Incorrect signal value description (invalid format)", expanded=True
        ):
            st.error(f"Found {len(invalid_val)} value descriptions with wrong format")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Signal Name": invalid_val.keys(),
                        "Incorrect Description": invalid_val.values(),
                    }
                )
            )
            st.info(
                "Allowed format examples:\n"
                "0x0: No Error\n"
                "0x0: <50% Alarm 0x1: <10% Alarm\n"
                "0x0~0x3: Reserved\n"
                "0x0: ACC_Off 0x1: ACC_Active\n"
                "0x0: 0% 0x1: 10%\n"
                "0x0: -8 level 0x1: -7 level\n"
                "0x0: Level 1(low) 0x1: Level 2(medium)\n"
                "0x0: AC Plug&DC Plug Connected"
            )

    if invalid_nan:
        with st.expander("NaN Value Description", expanded=True):
            st.error(f"Found {len(invalid_nan)} signals without value description")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Signal Name": invalid_nan.keys(),
                        "Description Status": invalid_nan.values(),
                    }
                )
            )

    return False if (invalid_nan or invalid_val or invalid_chars) else True


def validate_signal_descriprion(data_frame: pd.DataFrame) -> bool:
    sig_desc = dict(zip(data_frame["Sig Name"], data_frame["Description"]))

    invalid_val = {}
    invalid_nan = {}

    for mes, val in sig_desc.items():
        if pd.isna(val):
            invalid_nan[mes] = "NaN"
            continue

        str_val = str(val)
        if not re.fullmatch(r"^[A-Za-z0-9 ,.;:+_/-<>%()~-]+$", str_val):
            invalid_val[mes] = str_val

    if not invalid_nan and not invalid_val:
        st.success("All Signal Description are correct!")
        return True

    if invalid_val:
        with st.expander(
            "Incorrect signal description (not needed characters)", expanded=True
        ):
            st.error(
                f"Found {len(invalid_val)} value descriptions with wrong characters"
            )
            st.dataframe(
                pd.DataFrame(
                    {
                        "Signal Name": invalid_val.keys(),
                        "Incorrect Description": invalid_val.values(),
                    }
                )
            )
            st.info(
                "Allowed characters: A-Z, a-z, 0-9, spaces, commas, periods, and semicolons"
            )

    if invalid_nan:
        with st.expander("NaN Value Description", expanded=True):
            st.error(f"Found {len(invalid_nan)} signals without value description")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Signal Name": invalid_nan.keys(),
                        "Description Status": invalid_nan.values(),
                    }
                )
            )

    return False if (invalid_nan or invalid_val) else True


def validate_byte_order(data_frame: pd.DataFrame) -> bool:

    byte_order = dict(zip(data_frame["Sig Name"], data_frame["Byte Order"]))

    invalid_order = {}

    for mes, byte in byte_order.items():
        if byte != "Motorola MSB":
            invalid_order[mes] = byte

    if not invalid_order:
        st.success("All Signal Byte Orders are correct!")

    if invalid_order:
        with st.expander("Incorrect Byte Order", expanded=True):
            st.error(f"Found {len(invalid_order.keys())} incorrect byte order")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Signal Name": invalid_order.keys(),
                        "Incorrect Byte Order": invalid_order.values(),
                    }
                )
            )
            st.info("Byte Order in valid value 'Motorola MSB'")

    return False


def validate_start_byte(data_frame: pd.DataFrame) -> bool:
    start_byte = dict(zip(data_frame["Sig Name"], data_frame["Start Byte"]))

    invalid_byte = {}

    for sig, byte in start_byte.items():
        if byte not in range(0, 8):
            invalid_byte[sig] = byte

    if not invalid_byte:
        st.success("All Start Byte are correct!")
        return True

    if invalid_byte:
        with st.expander("Inccorect Start Byte", expanded=True):
            st.error(f"Found {len(invalid_byte.keys())} incorrect start byte")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Signal Name": invalid_byte.keys(),
                        "Incorrect Start Byte": invalid_byte.values(),
                    }
                )
            )
            st.info("Start Byte is only a number, in the range from 0 to 7")

    return False


def validate_start_bit(data_frame: pd.DataFrame) -> bool:
    start_bit = dict(zip(data_frame["Sig Name"], data_frame["Start Bit"]))

    invalid_bit = {}

    for sig, byte in start_bit.items():
        if byte not in range(0, 64):
            invalid_bit[sig] = byte

    if not invalid_bit:
        st.success("All Start Bit are correct!")
        return True

    if invalid_bit:
        with st.expander("Inccorect Start Bit", expanded=True):
            st.error(f"Found {len(invalid_bit.keys())} incorrect start bit")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Signal Name": invalid_bit.keys(),
                        "Incorrect Start Bit": invalid_bit.values(),
                    }
                )
            )
            st.info("Start Bit is only a number, in the range from 0 to 63")

    return False


def validate_signal_send_type(data_frame: pd.DataFrame) -> bool:
    sig_send_type = dict(zip(data_frame["Sig Name"], data_frame["Signal Send Type"]))
    msg_send_type = dict(zip(data_frame["Msg Name"], data_frame["Send Type"]))

    validation_rules = {
        "CA": ["Cycle", "IfActiveWithRepetition"],
        "CE": [
            "Cycle",
            "OnWrite",
            "OnChange",
            "OnWriteWithRepetition",
            "OnChangeWithRepetition",
        ],
        "Cycle": ["Cycle"],
        "Event": [
            "OnWrite",
            "OnChange",
            "OnWriteWithRepetition",
            "OnChangeWithRepetition",
        ],
        "IfActive": ["IfActive"],
    }

    invalid_signals = []

    for sig_name, sig_type in sig_send_type.items():
        msg_name = data_frame[data_frame["Sig Name"] == sig_name]["Msg Name"].iloc[0]
        msg_type = msg_send_type.get(msg_name)

        if msg_type in validation_rules:
            allowed_types = validation_rules[msg_type]
            if sig_type not in allowed_types:
                invalid_signals.append(
                    {
                        "Signal Name": sig_name,
                        "Message Name": msg_name,
                        "Message Send Type": msg_type,
                        "Signal Send Type": sig_type,
                        "Expected Types": ", ".join(allowed_types),
                    }
                )

    if not invalid_signals:
        st.success("All Signal Send Types are correct!")
        return True
    else:
        with st.expander("Invalid Signal Send Types", expanded=True):
            st.error(f"Found {len(invalid_signals)} invalid signal send types:")
            st.dataframe(pd.DataFrame(invalid_signals))
            st.info(
                """
            Validation rules:
            - If Msg Send Type == 'CA': Signal Send Type must be in ['Cycle', 'IfActiveWithRepetition']
            - If Msg Send Type == 'CE': Signal Send Type must be in ['Cycle', 'OnWrite', 'OnChange', 'OnWriteWithRepetition', 'OnChangeWithRepetition']
            - If Msg Send Type == 'Cycle': Signal Send Type must be 'Cycle'
            - If Msg Send Type == 'Event': Signal Send Type must be in ['OnWrite', 'OnChange', 'OnWriteWithRepetition', 'OnChangeWithRepetition']
            - If Msg Send Type == 'IfActive': Signal Send Type must be 'IfActive'
            """
            )
        return False


def validate_resolution(data_frame: pd.DataFrame) -> bool:
    resol = dict(zip(data_frame["Sig Name"], data_frame["Resolution"]))

    invalid_type = {}
    invalid_val = {}

    for sig, res in resol.items():
        if pd.isna(res):
            invalid_val[sig] = res
            continue

        if type(res) not in [int, float]:
            invalid_type[sig] = res

    if not invalid_type and not invalid_val:
        st.success("All Resolutions are correct!")
        return True

    if invalid_type:
        with st.expander("Incorrect Resolution type", expanded=True):
            st.error(f"Found {len(invalid_type.keys())} inccorect resolution type")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Sig Name": invalid_type.keys(),
                        "Incorrect Resolution": invalid_type.values(),
                    }
                )
            )
            st.info("Resolution is only int or float")

    if invalid_val:
        with st.expander("Incorrect Resolution value", expanded=True):
            st.error(f"Found {len(invalid_val.keys())} incorrect resolution value")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Sig Name": invalid_val.keys(),
                        "Incorrect Resolution": invalid_val.values(),
                    }
                )
            )
            st.info("Resolution must be in matrix")

    return False


def validate_offset(data_frame: pd.DataFrame) -> bool:
    offset = dict(zip(data_frame["Sig Name"], data_frame["Offset"]))

    invalid_val = {}
    invalid_type = {}

    for sig, off in offset.items():
        if pd.isna(off):
            invalid_val[sig] = off
            continue

        if type(off) not in [int, float]:
            invalid_type[sig] = off

    if not invalid_val and not invalid_type:
        st.success("All Signals Offset are correct!")
        return True

    if invalid_type:
        with st.expander("Incorrect Offset type", expanded=True):
            st.error(f"Found {len(invalid_type.keys())} incorrect offset type")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Sig Name": invalid_type.keys(),
                        "Incorrect Offset": invalid_type.values(),
                    }
                )
            )
            st.info("Offset must be int or float")

    if invalid_val:
        with st.expander("Incorrect Offset value", expanded=True):
            st.error(f"Found {len(invalid_val.keys())} incorrect offset value")
            st.dataframe(
                pd.DataFrame(
                    {
                        "Sig Name": invalid_val.keys(),
                        "Incorrect Offset": invalid_val.values(),
                    }
                )
            )
            st.info("Offset must be in matrix")

    return False


def validate_minimum(data_frame: pd.DataFrame) -> bool:
    min_phys = dict(zip(data_frame["Sig Name"], data_frame["Min"]))
    min_hex = dict(zip(data_frame["Sig Name"], data_frame["Min Hex"]))
    resolutions = dict(zip(data_frame["Sig Name"], data_frame["Resolution"]))
    offsets = dict(zip(data_frame["Sig Name"], data_frame["Offset"]))

    invalid_signals = []

    for sig_name in min_phys.keys():
        phys = min_phys.get(sig_name)
        hex_val = min_hex.get(sig_name)
        res = resolutions.get(sig_name)
        offset = offsets.get(sig_name, 0)

        if pd.isna(phys) or pd.isna(hex_val) or pd.isna(res):
            continue

        try:
            if isinstance(hex_val, str):
                if hex_val.startswith("0x"):
                    hex_int = int(hex_val, 16)
                else:
                    hex_int = int(hex_val)
            else:
                hex_int = int(hex_val)

            calculated_phys = hex_int * res + offset

            if not math.isclose(calculated_phys, phys, rel_tol=1e-9):
                invalid_signals.append(
                    {
                        "Signal Name": sig_name,
                        "Min (Physical)": phys,
                        "Min (Hex)": hex_val,
                        "Calculated Physical": calculated_phys,
                        "Resolution": res,
                        "Offset": offset,
                        "Difference": abs(calculated_phys - phys),
                    }
                )

        except (ValueError, TypeError) as e:
            invalid_signals.append(
                {
                    "Signal Name": sig_name,
                    "Error": f"Invalid data format: {str(e)}",
                    "Min (Hex)": hex_val,
                    "Resolution": res,
                    "Offset": offset,
                }
            )

    if not invalid_signals:
        st.success(
            "All minimum values match the formula: Physical = (Hex * Resolution) + Offset"
        )
        return True
    else:
        with st.expander("Invalid Minimum Values", expanded=True):
            st.error(
                f"Found {len(invalid_signals)} signals with incorrect minimum values"
            )

            df_errors = pd.DataFrame(invalid_signals)

            if "Error" in df_errors.columns:
                st.warning("Some values have format issues:")
                st.dataframe(
                    df_errors[df_errors["Error"].notna()][["Signal Name", "Error"]]
                )

                df_errors = df_errors[df_errors["Error"].isna()]

            if not df_errors.empty:
                st.dataframe(df_errors)
                st.info("Physical value should equal (Hex * Resolution) + Offset")

        return False


def validate_maximum(data_frame: pd.DataFrame) -> bool:
    max_phys = dict(zip(data_frame["Sig Name"], data_frame["Max"]))
    max_hex = (
        dict(zip(data_frame["Sig Name"], data_frame["Max Hex"]))
        if "Max Hex" in data_frame.columns
        else {}
    )
    resolutions = dict(zip(data_frame["Sig Name"], data_frame["Resolution"]))
    offsets = dict(zip(data_frame["Sig Name"], data_frame["Offset"]))
    last_sign_val_desc = dict(zip(data_frame["Sig Name"], data_frame["Signal Value Description"]))

    invalid_sig_val_desc = {}
    error_with_desc = {}

    for sig_name, val in last_sign_val_desc.items():
        if pd.isna(val):
            invalid_sig_val_desc[sig_name] = "NaN"
            continue
        
        str_val = str(val).strip().split("\n")[-1].split(":")[0].split("~")[-1]
        try:
            num = int(str_val, 16)
            if max_phys[sig_name] < num:
                error_with_desc[sig_name] = {
                    "Max Val Phys": max_phys[sig_name],
                    "Max Val Description": num
                }

        except ValueError:
            st.error(f"Cannot convert str {str_val}, to int")


    if not max_hex:
        max_hex = dict(zip(data_frame["Sig Name"], data_frame["Invalid"]))

    invalid_signals = []

    for sig_name in max_phys.keys():
        phys = max_phys.get(sig_name)
        hex_val = max_hex.get(sig_name)
        res = resolutions.get(sig_name)
        offset = offsets.get(sig_name, 0)

        if pd.isna(phys) or pd.isna(hex_val) or pd.isna(res):
            continue

        try:
            if isinstance(hex_val, str):
                if hex_val.startswith("0x"):
                    hex_int = int(hex_val, 16)
                else:
                    hex_int = int(hex_val)
            else:
                hex_int = int(hex_val)

            calculated_phys = hex_int * res + offset
            difference = abs(calculated_phys - phys)
            
            if not math.isclose(calculated_phys, phys, rel_tol=1e-9) and difference >= 1:
                invalid_signals.append(
                    {
                        "Signal Name": sig_name,
                        "Max (Physical)": phys,
                        "Max (Hex)": hex_val,
                        "Calculated Physical": calculated_phys,
                        "Resolution": res,
                        "Offset": offset,
                        "Difference": difference,
                    }
                )

        except (ValueError, TypeError) as e:
            invalid_signals.append(
                {
                    "Signal Name": sig_name,
                    "Error": f"Invalid data format: {str(e)}",
                    "Max (Hex)": hex_val,
                    "Resolution": res,
                    "Offset": offset,
                }
            )

    if not invalid_signals and not error_with_desc:
        st.success(
            "All maximum values match the formula: Physical = (Hex * Resolution) + Offset"
        )
        return True
    else:
        with st.expander("Invalid Maximum Values", expanded=True):
            st.error(
                f"Found {len(invalid_signals)} signals with incorrect maximum values"
            )

            df_errors = pd.DataFrame(invalid_signals)

            if "Error" in df_errors.columns:
                st.warning("Some values have format issues:")
                st.dataframe(
                    df_errors[df_errors["Error"].notna()][["Signal Name", "Error"]]
                )

                df_errors = df_errors[df_errors["Error"].isna()]

            if not df_errors.empty:
                st.dataframe(df_errors)
                st.info("Physical value should equal (Hex * Resolution) + Offset")
        
        with st.expander("Invalid Max Phys and Max Description", expanded=True):
            st.error(f"Found {len(error_with_desc.keys())} signals, which have Phys value less which Signal Value Description")
            data = {
                "Signal Name": list(error_with_desc.keys()),
                "Max Value Phys": [v["Max Val Phys"] for v in error_with_desc.values()],
                "Signal Value Description": [v["Max Val Description"] for v in error_with_desc.values()]
            }
            st.dataframe(pd.DataFrame(data))
            st.info("The maximum Phys/Hex value must be greater than or equal to the last value in the signal value description.")
        return False


def validate_cycle_times(data_frame):
    errors = []

    for idx, row in data_frame.iterrows():
        cycle_time = row["Cycle Type"]
        send_type = row["Send Type"]

        if send_type in ["Cycle", "CE", "CA"]:
            if pd.isna(cycle_time):
                errors.append(
                    {
                        "Error Type": "Not Cycle Type",
                        "Message": row["Msg Name"],
                        "Send Type": send_type,
                        "Value": "Empty",
                    }
                )
            elif not isinstance(cycle_time, (int, float)) or cycle_time < 0:
                errors.append(
                    {
                        "Error Type": "Некорректное время",
                        "Сообщение": row["Msg Name"],
                        "Тип отправки": send_type,
                        "Значение": cycle_time,
                    }
                )

        elif send_type in ["Event", "IfActive"]:
            if not pd.isna(cycle_time):
                errors.append(
                    {
                        "Error Type": "Лишнее время цикла",
                        "Сообщение": row["Msg Name"],
                        "Тип отправки": send_type,
                        "Значение": cycle_time,
                    }
                )

    if not errors:
        st.success("Проверка времени цикла выполнена успешно!")
        return True

    error_df = pd.DataFrame(errors)
    with st.expander("Ошибки во времени цикла", expanded=True):
        st.error(f"Найдено {len(error_df)} ошибок")
        st.dataframe(error_df)
        st.info("Для Cycle/CE/CA - обязательно, для Event/IfActive - должно быть пусто")

    return False

# Initinal
def validate_signal_values_against_bit_length(data_frame: pd.DataFrame) -> bool:
    signal_lengths = dict(zip(data_frame["Sig Name"], data_frame["Length"].astype(int)))
    max_values = dict(zip(data_frame["Sig Name"], data_frame["Max"]))
    
    try:
        initial_values = dict(zip(
            data_frame["Sig Name"], 
            data_frame["Initinal"].apply(
                lambda x: int(x, 16) if isinstance(x, str) and x.startswith(("0x", "0X")) 
                else int(x) if isinstance(x, str) 
                else x
            )
        ))
        
        invalid_values = dict(zip(
            data_frame["Sig Name"], 
            data_frame["Invalid"].apply(
                lambda x: int(x, 16) if isinstance(x, str) and x.startswith(("0x", "0X")) 
                else int(x) if isinstance(x, str) 
                else x
            )
        ))
    except (ValueError, AttributeError) as e:
        st.error(f"Failed to parse Initial or Invalid values: {str(e)}")
        return False

    invalid_signals = []

    for sig_name, length in signal_lengths.items():
        if pd.isna(length):
            continue
            
        max_allowed = (1 << int(length)) - 1
        
        max_val = max_values.get(sig_name)
        init_val = initial_values.get(sig_name)
        inval_val = invalid_values.get(sig_name)

        if not pd.isna(max_val) and max_val > max_allowed:
            invalid_signals.append({
                "Signal Name": sig_name,
                "Value Type": "Max",
                "Value": max_val,
                "Bit Length": length,
                "Max Allowed": max_allowed
            })

        if not pd.isna(init_val) and init_val > max_allowed:
            invalid_signals.append({
                "Signal Name": sig_name,
                "Value Type": "Initial",
                "Value": init_val,
                "Bit Length": length,
                "Max Allowed": max_allowed
            })

        if not pd.isna(inval_val) and inval_val > max_allowed:
            invalid_signals.append({
                "Signal Name": sig_name,
                "Value Type": "Invalid",
                "Value": inval_val,
                "Bit Length": length,
                "Max Allowed": max_allowed
            })

    if not invalid_signals:
        st.success("All signal values are within bit length limits!")
        return True
    else:
        with st.expander("Signal Values Exceeding Bit Length Limits", expanded=True):
            st.error(f"Found {len(invalid_signals)} signals exceeding bit length limits")
            st.dataframe(pd.DataFrame(invalid_signals))
            st.info(
                "Signal values (Max, Initial, Invalid) must not exceed 2^N - 1, "
                "where N is the signal bit length"
            )
        return False



def main():
    st.title("🚧CAN Messages Validator")
    uploaded_file = st.file_uploader("Upload matrix file", type=["xlsx"])

    if uploaded_file:
        try:
            df = load_xlsx(uploaded_file)
            processed_df = create_correct_df(df)
            file_attr = get_file_info(uploaded_file.name)
            st.success("File loaded successfully!")

            if st.button("Export All Validation Errors to Excel"):
                output_path = f"{file_attr['protocol']}_{file_attr['domain_name']}_{file_attr['date']}_highlighted_errors_{datetime.now().strftime('%Y%m%d')}.xlsx"
                if export_validation_errors_to_excel(processed_df, uploaded_file, output_path):
                    st.success(f"Validation errors highlighted in {output_path}")
                    with open(output_path, "rb") as f:
                        st.download_button(
                            label="Download Highlighted File",
                            data=f,
                            file_name=output_path,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                else:
                    st.success("No validation errors found!")

            (
                tab1,
                tab2,
                tab3,
                tab4,
                tab5,
                tab6,
                tab7,
                tab8,
                tab9,
                tab10,
                tab11,
                tab12,
                tab13,
                tab14,
                tab15,
                tab16,
                tab17,
                tab18,
                tab20
                # tab19
            ) = st.tabs(
                [
                    "Message Names",
                    "Message Types",
                    "Messages IDs",
                    "Messages Send Type",
                    "Messages Frame Format",
                    "Messages BRS",
                    "Messages Lenght",
                    "Signal Name",
                    "Signal Value Description",
                    "Signal Description",
                    "Byte Order",
                    "Start Byte",
                    "Start Bit",
                    "Signal Send Type",
                    "Resolution",
                    "Offset",
                    "Minimum",
                    "Maximum",
                    # "ECU Consistency"
                    "Signal Values Against Bit Length"
                ]
            )

            with tab1:
                validate_messages_name(processed_df)

            with tab2:
                validate_messages_type(processed_df)

            with tab3:
                validate_messages_id(processed_df)

            with tab4:
                validate_messages_send_type(processed_df)

            with tab5:
                validate_messages_frame_fromat(uploaded_file.name, processed_df)

            with tab6:
                validate_messages_BRS(uploaded_file.name, processed_df)

            with tab7:
                validate_messages_length(uploaded_file.name, processed_df)

            with tab8:
                validate_signal_names(processed_df)

            with tab9:
                validate_signal_value_description(processed_df)

            with tab10:
                validate_signal_descriprion(processed_df)

            with tab11:
                validate_byte_order(processed_df)

            with tab12:
                validate_start_byte(processed_df)

            with tab13:
                validate_start_bit(processed_df)

            with tab14:
                validate_signal_send_type(processed_df)

            with tab15:
                validate_resolution(processed_df)

            with tab16:
                validate_offset(processed_df)

            with tab17:
                validate_minimum(processed_df)

            with tab18:
                validate_maximum(processed_df)

            # with tab19:
            #     validate_cycle_times(processed_df)

            with tab20:
                validate_signal_values_against_bit_length(processed_df)

        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
    else:
        st.info("Please upload an Excel file to begin validation")


if __name__ == "__main__":
    main()
