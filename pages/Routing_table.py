import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from copy import copy
from datetime import datetime
from io import BytesIO
import json
import itertools


def set_page_config():
    st.title("üîÑRouting Table")


def gateway_selection():
    return st.radio("Choose gateway:", ("SGW", "CGW"))


def files_upload():
    uploaded_files = st.file_uploader(
        "Load domain matrices", type=["xlsx"], accept_multiple_files=True
    )
    if uploaded_files:
        st.success(f"‚úÖ Uploaded matrices: {len(uploaded_files)}")
        for file in uploaded_files:  # –æ–≥—Ä–∞–Ω–∏—á–∏–≤–∞–µ–º –≤—ã–≤–æ–¥
            st.write(file.name)
    return uploaded_files


def get_pd_data(uploaded_files):
    if uploaded_files:
        pd_df_matrices = {}
        # –î–ª—è –∫–∞–∂–¥–æ–≥–æ —Ñ–∞–π–ª–∞ –ø–æ–ª—É—á–∏—Ç—å –ø–∞—Ä—É: –∏–º—è - –¥–∞—Ç–∞—Ñ—Ä–µ–π–º
        for file in uploaded_files:
            df = pd.read_excel(file, sheet_name="Matrix")
            message_id_column_name = df.columns[2]
            df.dropna(subset=[message_id_column_name], inplace=True)
            pd_df_matrices[file.name] = df

        return pd_df_matrices
    return 0


def get_routing_table_template_path(input_path, output_path, uploaded_files):
    if uploaded_files:
        # –ó–∞–≥—Ä—É–∑–∏—Ç—å —Å—ã—Ä–æ–π —à–∞–±–ª–æ–Ω
        routing_table_template = load_workbook(input_path)
        with open(
            "./pages/template_values.json", "r", encoding="utf-8"
        ) as template_values_json:
            template_values = json.load(template_values_json)
        all_domains = {"BD", "DG", "PT", "CH", "DZ", "ET", "SGW"}
        routed_domains = []
        # –û–ø—Ä–µ–¥–µ–ª–∏—Ç—å –¥–æ–º–µ–Ω—ã –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã—Ö –º–∞—Ç—Ä–∏—Ü
        for file in uploaded_files:
            for domain in all_domains:
                if domain in file.name:
                    routed_domains.append(domain)

        real_values = {
            "release date": datetime.now().strftime("%Y.%m.%d"),
            "source domains": routed_domains,
            "target domains": routed_domains,
        }

        sheet_names = routing_table_template.sheetnames
        worksheets = {
            sheet_names[i]: routing_table_template.worksheets[i]
            for i in range(len(sheet_names))
        }
        # –ó–∞–ø–æ–ª–Ω–∏—Ç—å –ø–µ—Ä–≤—ã–π –ª–∏—Å—Ç (–æ–±–ª–æ–∂–∫–∞)
        for row in worksheets[sheet_names[0]].iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "Current date" in cell.value:
                        cell.value = cell.value.replace(
                            template_values["release date"], real_values["release date"]
                        )
        # –ó–∞–ø–æ–ª–Ω–∏—Ç—å –≤—Ç–æ—Ä–æ–π –ª–∏—Å—Ç (–∏—Å—Ç–æ—Ä–∏—è –∏–∑–º–µ–Ω–µ–Ω–∏–π)
        for row in worksheets[sheet_names[1]].iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    key = next(
                        (k for k, v in template_values.items() if v == cell.value), None
                    )
                    if key:
                        if isinstance(real_values[key], str):
                            cell.value = real_values[key]
                        else:
                            cell.value = ", ".join(real_values[key])

        routing_table_template.save(output_path)
        routing_table_template.close()

        return output_path
    return 0


def calculate_routing_table_data(pd_df_matrices, gateway):
    if pd_df_matrices and gateway:
        # –î–ª—è —Ö—Ä–∞–Ω–µ–Ω–∏—è –≤—Å–µ—Ö message_id –¥–ª—è –∫–∞–∂–¥–æ–π –º–∞—Ç–∞—Ä–∏—Ü—ã
        matrix_message_ids = {}
        # –î–ª—è —Ö—Ä–∞–Ω–µ–Ω–∏—è –æ–±—â–∏—Ö –º–µ–∂–¥—É –ø–∞—Ä–æ–π –º–∞—Ç—Ä–∏—Ü —Ñ—Ä–µ–π–º–æ–≤
        source_target_common_ids = {}
        # –î–ª—è —Ö—Ä–∞–Ω–µ–Ω–∏—è –º–∞—Ä—à—Ä—É—Ç–∏–∑–∏—Ä—É–µ–º—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π –¥–ª—è –∫–∞–∂–¥–æ–π –ø–∞—Ä—ã –∏–º–µ–Ω –º–∞—Ç—Ä–∏—Ü
        routing_table_data = {}
        # –ü–æ–ª—É—á–µ–Ω–∏–µ id –≤—Å–µ—Ö —Å–æ–æ–±—â–µ–Ω–∏–π —Ä–∞—Å—Å–º–∞—Ç—Ä–∏–≤–∞–µ–º–æ–π –º–∞—Ç—Ä–∏—Ü—ã
        for matrix_name, pd_df_matrix in pd_df_matrices.items():
            matrix_message_ids[matrix_name] = pd_df_matrix.iloc[:, 2]
        # –ü–æ–ª—É—á–µ–Ω–∏–µ –≤—Å–µ—Ö –ø–∞—Ä –º–∞—Ç—Ä–∏—Ü –¥–ª—è –ø–æ—Å—Ç—Ä–æ–µ–Ω–∏—è —Ç–∞–±–ª–∏—Ü—ã –º–∞—Ä—à—Ä—É—Ç–æ–≤
        source_target_data = list(itertools.permutations(pd_df_matrices.keys(), 2))
        for source_taget in source_target_data:
            source = source_taget[0]
            target = source_taget[1]
            # –ü–æ–ª—É—á–µ–Ω–∏–µ –æ–±—â–∏—Ö —Å–æ–æ–±—â–µ–Ω–∏–π –¥–ª—è —Ä–∞—Å—Å–º–∞—Ç—Ä–∏–≤–∞–µ–º–æ–π –ø–∞—Ä—ã –º–∞—Ç—Ä–∏—Ü
            routed_ids = pd.merge(
                matrix_message_ids[source], matrix_message_ids[target], how="inner"
            )
            # –î–ª—è –∫–∞–∂–¥–æ–π –ø–∞—Ä—ã –º–∞—Ç—Ç—Ä–∏—Ü –æ–ø—Ä–µ–¥–µ–ª—è—é—Ç—Å—è id –º–∞—Ä—à—Ä—É—Ç–∏–∑–∏—Ä—É–µ–º—ã—Ö —Å–æ–æ–±—â–µ–Ω–∏–π
            source_target_common_ids[source_taget] = routed_ids
            # –û–ø—Ä–µ–¥–µ–ª—è—é—Ç—Å—è —Å–æ–æ–±—â–µ–Ω–∏—è –º–∞—Ä—à—Ä—É—Ç–∏–∑–∏—Ä—É–µ–º—ã–µ –∏–∑ –º–∞—Ç—Ä–∏—Ü—ã –∏—Å—Ç–æ—á–Ω–∏–∫–∞
            source_matrix = pd_df_matrices[source]
            message_id_column_name = source_matrix.columns[2]
            source_gateway_column = source_matrix.filter(like=gateway).columns
            # –ü—Ä–æ–≤–µ—Ä–∫–∞ –Ω–∞ –Ω–∞–ª–∏—á–∏–µ –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ —à–ª—é–∑–∞ –≤ ECU –º–∞—Ç—Ä–∏—Ü—ã –∏—Å—Ç–æ—á–Ω–∏–∫–∞
            if len(source_gateway_column) == 0:
                st.error(f"'{gateway}' not in '{source}'. Check gateway.")
                st.stop()
            routed_ids_list = routed_ids[message_id_column_name].tolist()
            source_matrix_routed_messages = source_matrix[
                (source_matrix[message_id_column_name].isin(routed_ids_list))
                & (source_matrix[source_gateway_column[0]] == "R")
            ]
            # –û–ø—Ä–µ–¥–µ–ª—è—é—Ç—Å—è —Å–æ–æ–±—â–µ–Ω–∏—è –º–∞—Ä—à—Ä—É—Ç–∏–∑–∏—Ä—É–µ–º—ã–µ –≤ —Ü–µ–ª–µ–≤—É—é –º–∞—Ç—Ä–∏—Ü—É
            target_matrix = pd_df_matrices[target]
            message_id_column_name = target_matrix.columns[2]
            target_gateway_column = target_matrix.filter(like=gateway).columns
            # –ü—Ä–æ–≤–µ—Ä–∫–∞ –Ω–∞ –Ω–∞–ª–∏—á–∏–µ –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ —à–ª—é–∑–∞ –≤ ECU –º–∞—Ç—Ä–∏—Ü—ã –ø–æ–ª—É—á–∞—Ç–µ–ª—è
            if len(target_gateway_column) == 0:
                st.error(f"'{gateway}' not in '{target}'. Check gateway.")
                st.stop()
            target_matrix_routed_messages = target_matrix[
                (target_matrix[message_id_column_name].isin(routed_ids_list))
                & (target_matrix[target_gateway_column[0]] == "S")
            ]
            # –û–ø—Ä–µ–¥–µ–ª—è—é—Ç—Å—è –º–∞—Ä—à—Ä—É—Ç–∏–∑–∏—Ä—É–µ–º—ã–µ —Å–æ–æ–±—â–µ–Ω–∏—è –º–µ–∂–¥—É –¥–∞–Ω–Ω—ã–º–∏ –º–∞—Ç—Ä–∏—Ü–∞–º–∏
            matrices_routed_messages = pd.merge(
                source_matrix_routed_messages,
                target_matrix_routed_messages,
                how="inner",
            )
            matrices_routed_messages = matrices_routed_messages[
                [
                    matrices_routed_messages.columns[0],
                    matrices_routed_messages.columns[2],
                ]
            ]
            # –ó–∞–ø–∏—Å—å –≤ —Å–ª–æ–≤–∞—Ä—å –º–∞—Ä—à—Ä—É—Ç–∏–∑–∏—Ä—É–µ–º—ã—Ö –º–µ–∂–¥—É –¥–∞–Ω–Ω—ã–º–∏ –º–∞—Ç—Ä–∏—Ü–∞–º–∏ —Å–æ–æ–±—â–µ–Ω–∏–π –ø–æ –∫–ª—é—á—É —Ä–∞—Å—Å–º–∞—Ç—Ä–∏–≤–∞–µ–º–æ–π –ø–∞—Ä—ã –º–∞—Ç—Ä–∏—Ü
            routing_table_data[source_taget] = matrices_routed_messages

        return routing_table_data
    return 0


def generate_routing_table(routing_table_data, routing_table_template_path, gateway):
    if (routing_table_data) and (routing_table_template_path) and (gateway):
        generate_btn = st.button("Generate")
        if generate_btn:
            # –ó–∞–≥—Ä—É–∑–∏—Ç—å –ø–æ–¥–≥–æ—Ç–æ–≤–ª–µ–Ω–Ω—ã–π —à–∞–±–ª–æ–Ω
            routing_table = load_workbook(routing_table_template_path)
            table_headers = [
                "Signal Name",
                "Message Name",
                "Message ID",
                "Signal Name",
                "Message Name",
                "Message ID",
                "Routing Type",
                "Gateway ECU",
                "Change Record",
            ]
            start_row = 3
            start_col = 1
            merge_count = 3
            route_table_worksheet = routing_table.worksheets[-1]
            current_col = start_col
            current_raw = start_row
            # –Ω–∞—á–∞–ª–æ —Å—Ç–∏–ª–∏ (–Ω—É–∂–Ω–æ –ø–µ—Ä–µ–Ω–µ—Å—Ç–∏ –≤ –∏–∑–æ–ª–∏—Ä–æ–≤–∞–Ω–Ω–æ–µ –º–µ—Å—Ç–æ)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            blue_fill = PatternFill(start_color="00ccff", fill_type="solid")
            orange_fill = PatternFill(start_color="ff9900", fill_type="solid")
            yellow_fill = PatternFill(start_color="ffff99", fill_type="solid")
            green_fill = PatternFill(start_color="ccffcc", fill_type="solid")

            text_center_alignment = Alignment(horizontal="center", vertical="center")

            custom_font_bold = Font(
                name="Calibri",
                size=12,
                bold=True,
            )

            custom_font = Font(
                name="Calibri",
                size=12,
            )
            # –∫–æ–Ω–µ—Ü —Å—Ç–∏–ª–∏ (–Ω—É–∂–Ω–æ –ø–µ—Ä–µ–Ω–µ—Å—Ç–∏ –≤ –∏–∑–æ–ª–∏—Ä–æ–≤–∞–Ω–Ω–æ–µ –º–µ—Å—Ç–æ)
            for matrix_pair, route_data in routing_table_data.items():
                # –û–ø—Ä–µ–¥–µ–ª–∏—Ç—å –¥–∞–Ω–Ω—ã–µ —Å—Ç—Ä–æ–∫–∏ —Å –∏—Å—Ç–æ—á–Ω–∏–∫–æ–º –∏ –ø–æ–ª—É—á–∞—Ç–µ–ª–µ–º
                source_taget_header = [
                    f"Source: {matrix_pair[0]}",
                    f"Target: {matrix_pair[1]}",
                    None,
                ]
                # –ó–∞–ø–æ–ª–Ω–∏—Ç—å —Å—Ç—Ä–æ–∫—É —Å –∏—Å—Ç–æ—á–Ω–∏–∫–æ–º –∏ –ø–æ–ª—É—á–∞—Ç–µ–ª–µ–º
                for i in range(merge_count):
                    cell = route_table_worksheet.cell(
                        row=current_raw,
                        column=current_col,
                        value=source_taget_header[i],
                    )
                    cell.border = thin_border
                    cell.font = custom_font_bold
                    cell.alignment = text_center_alignment
                    route_table_worksheet.merge_cells(
                        start_row=current_raw,
                        start_column=current_col,
                        end_row=current_raw,
                        end_column=current_col + merge_count - 1,
                    )
                    current_col += merge_count

                current_raw += 1
                current_col = start_col
                # –ó–∞–ø–æ–ª–Ω–∏—Ç—å –∑–∞–≥–æ–ª–æ–≤–∫–∏ —Ç–∞–±–ª–∏—Ü—ã –º–∞—Ä—à—Ä—É—Ç–∏–∑–∞—Ü–∏–∏ –¥–ª—è —Ä–∞—Å—Å–º–∞—Ç—Ä–∏–≤–∞–µ–º–æ–π –ø–∞—Ä—ã
                for table_header in table_headers:
                    cell = route_table_worksheet.cell(
                        row=current_raw, column=current_col, value=table_header
                    )
                    cell.border = thin_border
                    cell.font = custom_font
                    cell.fill = blue_fill
                    cell.alignment = text_center_alignment
                    current_col += 1

                current_raw += 1
                current_col = start_col
                # –ü–µ—Ä–µ–≤–µ—Å—Ç–∏ –¥–∞—Ç–∞—Ñ—Ä–µ–π–º –≤ —Å—Ç—Ä–æ–∫–∏
                for route_table_row_data in dataframe_to_rows(
                    route_data, index=False, header=False
                ):
                    # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ø–æ–ª–µ–π Message Name –∏ Message ID –æ—Ç–ø—Ä–∞–≤–∏—Ç–µ–ª—è –∏ –ø–æ–ª—É—á–∞—Ç–µ–ª—è
                    for cell_value in route_table_row_data:
                        # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ø–æ–ª–µ–π Message Name –∏ Message ID –æ—Ç–ø—Ä–∞–≤–∏—Ç–µ–ª—è
                        cell = route_table_worksheet.cell(
                            row=current_raw, column=current_col + 1, value=cell_value
                        )
                        cell.border = thin_border
                        cell.font = custom_font
                        cell.fill = green_fill
                        cell.alignment = text_center_alignment
                        # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ø–æ–ª–µ–π Message Name –∏ Message ID –ø–æ–ª—É—á–∞—Ç–µ–ª—è
                        cell = route_table_worksheet.cell(
                            row=current_raw, column=current_col + 4, value=cell_value
                        )
                        cell.border = thin_border
                        cell.font = custom_font
                        cell.fill = green_fill
                        cell.alignment = text_center_alignment
                        current_col += 1
                    current_col = start_col
                    # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ø–æ–ª—è Signal Name –æ—Ç–ø—Ä–∞–≤–∏—Ç–µ–ª—è
                    cell = route_table_worksheet.cell(
                        row=current_raw, column=current_col, value=None
                    )
                    cell.border = thin_border
                    cell.font = custom_font
                    cell.fill = orange_fill
                    cell.alignment = text_center_alignment
                    # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ø–æ–ª—è Signal Name –ø–æ–ª—É—á–∞—Ç–µ–ª—è
                    cell = route_table_worksheet.cell(
                        row=current_raw, column=current_col + 3, value=None
                    )
                    cell.border = thin_border
                    cell.font = custom_font
                    cell.fill = orange_fill
                    cell.alignment = text_center_alignment
                    # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ø–æ–ª—è Routing Type
                    cell = route_table_worksheet.cell(
                        row=current_raw, column=current_col + 6, value="Message"
                    )
                    cell.border = thin_border
                    cell.font = custom_font
                    cell.fill = yellow_fill
                    cell.alignment = text_center_alignment
                    # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ø–æ–ª—è Gateway ECU
                    cell = route_table_worksheet.cell(
                        row=current_raw, column=current_col + 7, value=gateway
                    )
                    cell.border = thin_border
                    cell.font = custom_font
                    cell.fill = yellow_fill
                    cell.alignment = text_center_alignment
                    # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –ø–æ–ª—è Change Record
                    cell = route_table_worksheet.cell(
                        row=current_raw, column=current_col + 8, value=None
                    )
                    cell.border = thin_border
                    cell.font = custom_font
                    cell.fill = yellow_fill
                    cell.alignment = text_center_alignment
                    # –ø–µ—Ä–µ—Ö–æ–¥ –Ω–∞ —Å–ª–µ–¥—É—é—â—É—é —Å—Ç—Ä–æ–∫—É
                    current_col = start_col
                    current_raw += 1

                current_raw += 1

            # routing_table.save("routing_table.xlsx")
            routing_table.close()
        else:
            routing_table = 0

        return routing_table
    return 0


def download_routing_table(routing_table, gateway):
    if routing_table:
        output_path = (
            f"RoutingMAP-{gateway}_VNone_{datetime.now().strftime("%Y%m%d")}.xlsx"
        )
        buffer = BytesIO()
        routing_table.save(buffer)
        buffer.seek(0)
        st.download_button(
            label="Download routing map",
            data=buffer,
            file_name=output_path,
            mime="application/octet-stream",
            type="primary",
            help="Press to download .xlsx RoutingMap",
        )


def main():
    try:
        # –û–∑–∞–≥–ª–∞–≤–∏—Ç—å —Å—Ç—Ä–∞–Ω–∏—Ü—É
        set_page_config()
        # –ü—Ä–µ–¥–æ—Å—Ç–∞–≤–∏—Ç—å —Ñ–æ—Ä–º—É –¥–ª—è –∑–∞–≥—Ä—É–∑–∫–∏ —Ñ–∞–π–ª–æ–≤
        uploaded_files = files_upload()
        # –ü—Ä–µ–¥–æ—Å—Ç–∞–≤–∏—Ç—å –≤—ã–±–æ—Ä —à–ª—é–∑–∞
        gateway = gateway_selection()
        # –ü–æ–¥–≥–æ—Ç–æ–≤–∏—Ç—å —à–∞–±–ª–æ–Ω –∏ –ø–æ–ª—É—á–∏—Ç—å –µ–≥–æ –¥–∏—Ä–µ–∫—Ç–æ—Ä–∏—é
        routing_table_template_path = get_routing_table_template_path(
            "./pages/routing_table_template.xlsx",
            f"routing_table_template{datetime.now().strftime("%Y_%m_%d")}.xlsx",
            uploaded_files,
        )
        # –°—á–∏—Ç–∞—Ç—å –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã–µ —Ñ–∞–π–ª—ã –≤ pandas –¥–∞—Ç–∞—Ñ—Ä–µ–π–º—ã
        pd_df_matrices = get_pd_data(uploaded_files)
        # –û–±—Ä–∞–±–æ—Ç–∞—Ç—å –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã–µ –¥–∞–Ω–Ω—ã–µ –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –∑–∞–ø–æ–ª–Ω–µ–Ω–∏—è —Ç–∞–±–ª–∏—Ü—ã –º–∞—Ä—à—Ä—É—Ç–∏–∑–∞—Ü–∏–∏
        routing_table_data = calculate_routing_table_data(pd_df_matrices, gateway)
        # –ó–∞–ø–æ–ª–Ω–∏—Ç—å —Ç–∞–±–ª–∏—Ü—É –º–∞—Ä—à—Ä—É—Ç–∏–∑–∞—Ü–∏–∏ —Å —Ç—Ä–µ–±—É–µ–º—ã–º —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ–º
        routing_table = generate_routing_table(
            routing_table_data, routing_table_template_path, gateway
        )
        # –ü—Ä–µ–¥–æ—Å—Ç–∞–≤–∏—Ç—å –∫–æ–Ω—Ñ–∏–≥—É—Ä–∞—Ü–∏—é –¥–ª—è –∑–∞–≥—Ä—É–∑–∫–∏ –≤—ã—Ö–æ–¥–Ω–æ–≥–æ —Ñ–∞–π–ª–∞
        download_routing_table(routing_table, gateway)

    except Exception as e:
        st.error(f"Error occured: {str(e)}")
        st.stop()


if __name__ == "__main__":
    main()
