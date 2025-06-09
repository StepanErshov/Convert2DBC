import streamlit as st  # для браузера
import pandas as pd  # для работы с табличными данными
from openpyxl import load_workbook  # для загрузки Excel-книг
from copy import copy  # для копирования стилей ячеек
from io import BytesIO  # для работы с буфером в памяти
from datetime import datetime  # для получения текущей даты
from openpyxl.utils import get_column_letter  # для получения буквенных индексов столбцов
from openpyxl.styles import Alignment  # для задания выравнивания текста
import zipfile  # для создания zip-архива
 
# Настройка страницы Streamlit
st.set_page_config(
    page_title="ECU Splitter",
    page_icon="🏴",
    layout="centered",
    initial_sidebar_state="expanded"
)
 
st.title("Domain Excel Splitter")
st.image("excel_split_example.png", caption="Domain Excel → ECU Excel")
 
uploaded_file = st.file_uploader("🏴 Загрузите Excel файл", type=["xlsx"])
 
 
def set_column_widths(ws, widths):
    for col_idx, width in widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
 
 
def copy_row_with_style(src_row, dest_ws, dest_row_idx):
    for i, orig_cell in enumerate(src_row, 1):
        new_cell = dest_ws.cell(row=dest_row_idx, column=i, value=orig_cell.value)
        new_cell.font = copy(orig_cell.font)
        new_cell.fill = copy(orig_cell.fill)
        new_cell.border = copy(orig_cell.border)
        new_cell.alignment = copy(orig_cell.alignment)
 
 
def process_matrix_sheet(wb, ecu_col_indexes):
    ws_matrix = wb["Matrix"]
    wb.remove(ws_matrix)
    new_matrix_ws = wb.create_sheet("Matrix")
 
    # Устанавливаем кастомную ширину столбцов
    custom_widths = {
        1: 15, 2: 7.5, 3: 5, 4: 5, 5: 5, 6: 15, 7: 5, 8: 5, 9: 20, 10: 40,
        11: 12.5, 12: 5, 13: 5, 14: 20, 15: 5, 16: 10, 17: 5, 18: 5, 19: 5,
        20: 5, 21: 5, 22: 5, 23: 5, 24: 5, 25: 5, 26: 5, 27: 15, 28: 5, 29: 5,
        30: 5, 31: 10, 32: 5, 33: 5, 34: 5, 35: 5, 36: 5, 37: 5, 38: 5, 39: 5,
        40: 5, 41: 5, 42: 5, 43: 5, 44: 5, 45: 5, 46: 5,
    }
    set_column_widths(new_matrix_ws, custom_widths)
 
    # Копируем заголовки с форматами
    for col_idx, cell in enumerate(ws_matrix[1], 1):
        new_cell = new_matrix_ws.cell(row=1, column=col_idx, value=cell.value)
        new_cell.font = copy(cell.font)
        new_cell.fill = copy(cell.fill)
        new_cell.border = copy(cell.border)
        new_cell.alignment = copy(cell.alignment)
        if col_idx == 27:
            alignment = copy(cell.alignment)
            new_cell.alignment = Alignment(
                horizontal=alignment.horizontal,
                vertical=alignment.vertical,
                wrap_text=True
            )
 
    # Копируем строки, соответствующие ECU
    dest_row = 2
    for row in ws_matrix.iter_rows(min_row=2, max_row=ws_matrix.max_row):
        if any(
                (row[idx].value is not None and str(row[idx].value).lower() in ['s', 'r'])
                for idx in ecu_col_indexes.values()
        ):
            copy_row_with_style(row, new_matrix_ws, dest_row)
            dest_row += 1
 
    # Настройка переноса в столбце AA
    for row in new_matrix_ws.iter_rows(min_row=2, max_row=new_matrix_ws.max_row):
        aa_cell = row[26]
        aa_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
 
    # Группировка строк по полю A и I
    row_idx = 2
    while row_idx <= new_matrix_ws.max_row:
        a_val = new_matrix_ws.cell(row=row_idx, column=1).value
        if a_val:
            start_row = row_idx + 1
            end_row = start_row
            while end_row <= new_matrix_ws.max_row:
                i_val = new_matrix_ws.cell(row=end_row, column=9).value
                if i_val:
                    end_row += 1
                else:
                    break
            if end_row > start_row:
                for r in range(start_row, end_row):
                    new_matrix_ws.row_dimensions[r].outlineLevel = 1
                    new_matrix_ws.row_dimensions[r].hidden = True
                new_matrix_ws.row_dimensions[row_idx].collapsed = True
            row_idx = end_row
        else:
            row_idx += 1
 
    # Обрезаем лист
    max_rows, max_cols = 5000, 50
    if new_matrix_ws.max_row > max_rows:
        new_matrix_ws.delete_rows(max_rows + 1, new_matrix_ws.max_row - max_rows)
    if new_matrix_ws.max_column > max_cols:
        new_matrix_ws.delete_cols(max_cols + 1, new_matrix_ws.max_column - max_cols)
 
    return new_matrix_ws
 
 
def process_history_sheet(wb, ecu_col_indexes):
    if "History" not in wb.sheetnames:
        return False
    ws_history_orig = wb["History"]
    original_merged_ranges = list(ws_history_orig.merged_cells.ranges)
    new_history_ws = wb.copy_worksheet(ws_history_orig)
    wb.remove(ws_history_orig)
    new_history_ws.title = "History"
 
    for merged_range in original_merged_ranges:
        new_history_ws.merge_cells(str(merged_range))
 
    for col_idx, orig_cell in enumerate(ws_history_orig[1], 1):
        new_cell = new_history_ws.cell(row=1, column=col_idx, value=orig_cell.value)
        new_cell.font = copy(orig_cell.font)
        new_cell.fill = copy(orig_cell.fill)
        new_cell.border = copy(orig_cell.border)
        new_cell.alignment = copy(orig_cell.alignment)
 
    for row_idx in range(new_history_ws.max_row, 2, -1):
        cell_value = new_history_ws.cell(row=row_idx, column=6).value
        if cell_value not in ecu_col_indexes:
            new_history_ws.delete_rows(row_idx)
    return True
 
 
def identify_bus_users(df):
    """
    Identifies bus users (ECUs) by looking for columns that contain 'S' or 'R' values.
    Excludes the 'Unit\n单位' column from the results.
    """
    return [
        col
        for col in df.columns
        if any(val in ["S", "R"] for val in df[col].dropna().unique())
        and col != "Unit\n单位"
    ]
 
 
if uploaded_file:
    try:
        # Read the entire Matrix sheet to identify bus users
        df = pd.read_excel(uploaded_file, sheet_name="Matrix")
        bus_users = identify_bus_users(df)
         
        if not bus_users:
            st.error("🕭 Не найдены ECU в файле. Проверьте, что файл содержит столбцы с 'S' или 'R' значениями.")
            st.stop()
             
        # Create a container for checkboxes
        st.write("Выберите ECU для экспорта:")
        col1, col2, col3 = st.columns(3)  # Create 3 columns for better layout
        selected_ecus = []
 
        # Place checkboxes in columns
        for i, ecu in enumerate(bus_users):
            col = col1 if i % 3 == 0 else (col2 if i % 3 == 1 else col3)
            if col.checkbox(ecu, key=f"ecu_{ecu}"):
                selected_ecus.append(ecu)
 
        if selected_ecus and st.button("🖊 Экспорт выбранных ECU"):
            # Create a zip file in memory
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for selected_ecu in selected_ecus:
                    date_str = datetime.now().strftime("%d%m%Y")
                    ecu_list = [selected_ecu]
                    ecu_col_indexes = {ecu: df.columns.get_loc(ecu) for ecu in ecu_list}
 
                    # Create a new workbook for each ECU
                    wb = load_workbook(uploaded_file)
                    process_matrix_sheet(wb, ecu_col_indexes)
                    history_found = process_history_sheet(wb, ecu_col_indexes)
                    if not history_found:
                        st.warning(f"🕱 Лист 'History' не найден в ECU {selected_ecu}.")
 
                    # Save workbook to BytesIO
                    excel_buffer = BytesIO()
                    filename = f"ATOM_CAN_MATRIX_{selected_ecu}_{date_str}.xlsx"
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
 
                    # Add the Excel file to the zip
                    zip_file.writestr(filename, excel_buffer.getvalue())
 
            # Prepare zip file for download
            zip_buffer.seek(0)
            zip_filename = f"ECU_Export_{datetime.now().strftime('%d%m%Y_%H%M%S')}.zip"
 
            st.download_button(
                label=f"🖥 Скачать все файлы ({len(selected_ecus)} ECU)",
                data=zip_buffer,
                file_name=zip_filename,
                mime="application/zip",
                key="download_zip"
            )
            st.success(f"Экспорт завершён для {len(selected_ecus)} ECU. Все файлы готовы к скачиванию в одном архиве.")
    except Exception as e:
        st.error(f"🕭 Ошибка при обработке файла: {e}")