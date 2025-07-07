import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
from io import BytesIO
import re
import cantools

def set_page_config():
    st.title("üî•CAN ID Map")

def files_upload():
    uploaded_files = st.file_uploader("Load domain matrices or DBC", type=["xlsx", "dbc"], accept_multiple_files=True)
    uploaded_files_by_format = {}
    excel_files = []
    dbc_files = []
    if uploaded_files:
        st.success(f"Uploaded matrices: {len(uploaded_files)}")
        for file in uploaded_files:
            if file.name.endswith('.xlsx'):
                excel_files.append(file)
            else:
                dbc_files.append(file)
            st.write(file.name)
        uploaded_files_by_format["xlsx"] = excel_files
        uploaded_files_by_format["dbc"] = dbc_files

        return uploaded_files_by_format
    return 0

def get_format_splitted_files(uploaded_files):
    if uploaded_files:
        excel_files = uploaded_files["xlsx"]
        dbc_files = uploaded_files["dbc"]

        return excel_files, dbc_files
    return 0, 0

def input_version():
    version = st.text_input("Enter version (e.g., V3.1.1):")
    pattern = r"^V\d\.\d\.\d$"
    if version:
        if not re.match(pattern, version):
            st.error("Invalid format! Please enter as V1.2.3")
    else:
        version = "VNone"
    return version

def get_excel_2_df(uploaded_files):
    if uploaded_files:
        pd_df_matrices = []
        # –ü–æ–ª—É—á–∏—Ç—å –¥–∞—Ç–∞—Ñ—Ä–µ–π–º—ã –¥–ª—è –∫–∞–∂–¥–æ–≥–æ —Ñ–∞–π–ª–∞
        for file in uploaded_files:
            df = pd.read_excel(file, sheet_name="Matrix")
            necessary_columns = [0, 2, 4]
            df = df.iloc[:, necessary_columns]
            message_name_column = df.columns[0]
            df.dropna(subset=[message_name_column], inplace=True)
            df.columns = ['message name', 'message id', 'message cycle time']
            df['message id'] = df['message id'].str[2:]
            pd_df_matrices.append(df)

        return pd_df_matrices
    return 0

def get_dbc_2_df(uploaded_files):
    if uploaded_files:
        pd_df_matrices = []
        # –ü–æ–ª—É—á–∏—Ç—å –¥–∞—Ç–∞—Ñ—Ä–µ–π–º—ã –¥–ª—è –∫–∞–∂–¥–æ–≥–æ —Ñ–∞–π–ª–∞
        for file in uploaded_files:
            dbc_content = file.read().decode('utf-8')
            db = cantools.database.load_string(dbc_content, 'dbc')
            data = []
            for message in db.messages:
                message_id = hex(message.frame_id).upper()[2:]
                if len(message_id) == 2:
                    message_id = '0' + message_id
                data.append({
                    'message name': message.name,
                    'message id': message_id,
                    'message cycle time': message.cycle_time
                })
            df = pd.DataFrame(data)
            pd_df_matrices.append(df)

        return pd_df_matrices
    return 0

def get_merged_df(df_excel, df_dbc):
    if not isinstance(df_excel, int) and not isinstance(df_dbc, int):
        excel_dbc_df = df_excel + df_dbc
        merged_df = pd.concat(excel_dbc_df).drop_duplicates().reset_index(drop=True)
    elif not isinstance(df_excel, int):
        merged_df = pd.concat(df_excel).drop_duplicates().reset_index(drop=True)
    elif not isinstance(df_dbc, int):
        merged_df = pd.concat(df_dbc).drop_duplicates().reset_index(drop=True)
    else:
        return 0
    return merged_df

def stylise_cell(cell, msg_cycle_time):
    cycle_time_to_fill = {
        10.0:   PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red
        20.0:   PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),  # Orange/Gold
        50.0:   PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),  # Purple
        100.0:  PatternFill(start_color="33cc33", end_color="33cc33", fill_type="solid"),  # Green
        200.0:  PatternFill(start_color="66ff66", end_color="66ff66", fill_type="solid"),  # Light Green
        ">200.0": PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),  # Blue
        "nan":  PatternFill(start_color="FF66CC", end_color="FF66CC", fill_type="solid"),  # Pink
    }

    custom_font = Font(
        name='Arial',
        size=12,
    )

    cell.font = custom_font
    if float(msg_cycle_time) > 200:
        cell.fill = cycle_time_to_fill[">200.0"]
    elif pd.isna(msg_cycle_time):
        cell.fill = cycle_time_to_fill["nan"]
    else:
        cell.fill = cycle_time_to_fill[msg_cycle_time]

def get_check_result_ws(df, CAN_ID_Map):
    if CAN_ID_Map and not isinstance(df, int):
        id_map_ws = CAN_ID_Map["ATOM_ID Map"]
        check_result_ws = CAN_ID_Map.copy_worksheet(id_map_ws)
        check_result_ws.title = "CheckResult"

        return check_result_ws
    return 0

def get_overlays_df(df):
    if not isinstance(df, int):
        overlays_df = df[df['message id'].duplicated(keep=False)]

        return overlays_df
    return 0

def show_overlays(overlays_df):
    if not isinstance(overlays_df, int):
        if not overlays_df.empty:
            st.error("Overlayed ids:")
            st.write(overlays_df)

def get_multi_id_messages(df):
    if not isinstance(df, int):
        multi_id_messages = {}
        multi_id_messages_df = df[df['message name'].duplicated(keep=False)]
        multi_id_messages = multi_id_messages_df.groupby('message name')['message id'].unique().apply(list).to_dict()

        return multi_id_messages
    return 0

def show_multi_id_messages(multi_id_messages):
    if multi_id_messages:
        st.error("Multi ID messages:")
        st.write(multi_id_messages)

def generate_CAN_ID_Map(template_path, df):
    if template_path and not isinstance(df, int):
        # –ó–∞–≥—Ä—É–∑–∏—Ç—å —à–∞–±–ª–æ–Ω
        CAN_ID_Map = load_workbook(template_path)
        # –î–æ–±–∞–≤–∏—Ç—å –ª–∏—Å—Ç CheckResult
        check_result_ws = get_check_result_ws(df, CAN_ID_Map)
        # –ü–æ–ª—É—á–∏—Ç—å –¥–∞—Ç–∞—Ñ—Ä–µ–π–º —Å –Ω–∞–ª–æ–∂–µ–Ω–Ω—ã–º–∏ id
        overlays_df = get_overlays_df(df)
        # –í—ã–≤–µ—Å—Ç–∏ –≤ –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å –Ω–∞–ª–æ–∂–µ–Ω–Ω—ã–µ id
        show_overlays(overlays_df)
        # –ü–æ–ª—É—á–∏—Ç—å —Å–æ–æ–±—â–µ–Ω–∏—è —Å –Ω–µ–æ–¥–Ω–æ–∑–Ω–∞—á–Ω—ã–º–∏ id
        multi_id_messages = get_multi_id_messages(df)
        # –í—ã–≤–µ—Å—Ç–∏ –≤ –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å —Å–æ–æ–±—â–µ–Ω–∏—è —Å –Ω–µ–æ–¥–Ω–æ–∑–Ω–∞—á–Ω—ã–º–∏ id
        show_multi_id_messages(multi_id_messages)
        history_ws = CAN_ID_Map['History']
        id_map_ws = CAN_ID_Map["ATOM_ID Map"]
        # –ü–æ–¥–≥–æ—Ç–æ–≤–∏—Ç—å —Å–ª–æ–≤–∞—Ä—å –¥–ª—è –∫–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏–∏ msg_id –≤ id —Å—Ç–æ–ª–±—Ü–∞ excel 
        hex_column_id = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', 'A', 'B', 'C', 'D', 'E', 'F']
        excel_column_id =  ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']
        hex_to_column = dict(zip(hex_column_id, excel_column_id))
        # –ó–∞–ø–æ–ª–Ω–∏—Ç—å –¥–∞—Ç—É —Å–æ–∑–¥–∞–Ω–∏—è
        history_ws['B2'] = datetime.now().strftime("%d/%m/%Y")
        # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ —Ç–∞–±–ª–∏—Ü—ã
        for index, row in df.iterrows():
            msg_name = row.iloc[0]
            msg_id = row.iloc[1]
            msg_id_row = int(row.iloc[1][:2], 16)
            msg_id_column = hex_to_column[row.iloc[1][2:]]
            msg_cycle_time = row.iloc[2]
            # –ó–∞–ø–æ–ª–Ω–µ–Ω–∏–µ msg_name
            cell_id = f'{msg_id_column}{msg_id_row+1}'
            # –ü—Ä–æ–≤–µ—Ä–∫–∞ –Ω–∞ –Ω–∞–ª–æ–∂–µ–Ω–∏–µ –∏ –∑–∞–ø–æ–ª–Ω–µ–Ω–∏–µ —è—á–µ–µ–∫
            if msg_name in overlays_df.iloc[:, 0].values or msg_name in multi_id_messages:
                if not id_map_ws[cell_id].value:
                    id_map_ws[cell_id] = msg_name
                    check_result_ws[cell_id] = msg_name
                else:
                    id_map_ws[cell_id] = str(id_map_ws[cell_id].value) + f'\n{msg_name}'
                    check_result_ws[cell_id] = str(check_result_ws[cell_id].value) + f'\n{msg_name}'
            else:
                id_map_ws[cell_id] = msg_name
                stylise_cell(id_map_ws[cell_id], msg_cycle_time)

        CAN_ID_Map.save("CAN_ID_Map.xlsx")
        CAN_ID_Map.close()

        return CAN_ID_Map
    return 0

def download_CAN_ID_Map(CAN_ID_Map, version):
    if CAN_ID_Map:
        output_path = f"CANID Design_ATOM_{version}-{datetime.now().strftime("%Y%m%d")}.xlsx"
        buffer = BytesIO()
        CAN_ID_Map.save(buffer)
        buffer.seek(0)
        st.download_button(
            label="Download CAN ID map",
            data=buffer,
            file_name=output_path,
            mime="application/octet-stream",
            type="primary",
            help="Press to download .xlsx CAN ID map"
        )


def main():
    try:
        template_path = "pages/CAN_ID_Map_template.xlsx"

        # –û–∑–∞–≥–ª–∞–≤–∏—Ç—å —Å—Ç—Ä–∞–Ω–∏—Ü—É
        set_page_config()
        # –ü—Ä–µ–¥–æ—Å—Ç–∞–≤–∏—Ç—å —Ñ–æ—Ä–º—É –¥–ª—è –∑–∞–≥—Ä—É–∑–∫–∏ —Ñ–∞–π–ª–æ–≤
        uploaded_files = files_upload()
        # –ü–æ–ª—É—á–∏—Ç—å —Ñ–∞–π–ª—ã, —Ä–∞–∑–¥–µ–ª–µ–Ω–Ω—ã–µ –Ω–∞ xlsx –∏ dbc —Ñ–æ—Ä–º–∞—Ç—ã
        excel_files, dbc_files = get_format_splitted_files(uploaded_files)
        # –ü—Ä–µ–¥–æ—Å—Ç–∞–≤–∏—Ç—å –ø–æ–ª–µ –¥–ª—è –≤–≤–æ–¥–∞ –≤–µ—Ä—Å–∏–∏ —Ç–∞–±–ª–∏—Ü—ã
        version = input_version()
        # –ü–µ—Ä–µ–≤–µ—Å—Ç–∏ –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã–µ —Ç–∞–±–ª–∏—Ü—ã –≤ –¥–∞—Ç–∞—Ñ—Ä–µ–π–º
        df_excel = get_excel_2_df(excel_files)
        # –ü–µ—Ä–µ–≤–µ—Å—Ç–∏ –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã–µ DBC –≤ –¥–∞—Ç–∞—Ñ—Ä–µ–π–º
        df_dbc = get_dbc_2_df(dbc_files)
        # –°–æ–≤–º–µ—Å—Ç–∏—Ç—å excel –∏ dbc –¥–∞—Ç–∞—Ñ—Ä–µ–π–º—ã
        merged_df = get_merged_df(df_excel, df_dbc)
        # –°–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å id —Ç–∞–±–ª–∏—Ü—É
        CAN_ID_Map = generate_CAN_ID_Map(template_path, merged_df)
        # –°–∫–∞—á–∞—Ç—å —Ä–µ–∑—É–ª—å—Ç–∞—Ç
        download_CAN_ID_Map(CAN_ID_Map, version)


    except Exception as e:
        st.error(f"Error occured: {str(e)}")
        st.stop()

if __name__ == "__main__":
    main()