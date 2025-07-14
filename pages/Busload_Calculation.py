import streamlit as st
import pandas as pd
import cantools
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
from datetime import datetime
from io import BytesIO

def set_page_config():
    st.title("🚐 Busload Calculation")

def files_upload():
    uploaded_files = st.file_uploader("Load domain matrices or DBC", type=["xlsx", "dbc"], accept_multiple_files=True)
    uploaded_files_by_format = {}
    excel_files = []
    dbc_files = []
    if uploaded_files:
        st.success(f"Uploaded matrices: {len(uploaded_files)}")
        for file in uploaded_files:
            if file.name.endswith('.xlsx'):
                excel_files.append(file)
            else:
                dbc_files.append(file)
            st.write(file.name)
        uploaded_files_by_format["xlsx"] = excel_files
        uploaded_files_by_format["dbc"] = dbc_files

        return uploaded_files_by_format
    return 0

def input_version():
    version = st.text_input("Enter matrices release version:")
    if not version:
        version = "VNone"
    return version

def get_format_splitted_files(uploaded_files):
    if uploaded_files:
        if "xlsx" in uploaded_files:
            excel_files = uploaded_files["xlsx"]
        else:
            excel_files = []
        if "dbc" in uploaded_files:
            dbc_files = uploaded_files["dbc"]
        else:
            dbc_files = []

        return excel_files, dbc_files
    return 0, 0

def get_excel_2_df(excel_files):
    if excel_files:
        pd_df_matrices = {}
        # Получить датафреймы для каждого домена
        for file in excel_files:
            df = pd.read_excel(file, sheet_name="Matrix")
            if 'CANFD' in file.name:
                necessary_columns = [0, 2, 3, 4, 7]
            else:
                necessary_columns = [0, 2, 3, 4, 5]
            df = df.iloc[:, necessary_columns]
            message_name_column = df.columns[0]
            df.dropna(subset=[message_name_column], inplace=True)
            df.columns = ["Msg Name\n报文名称",	"Msg ID\n报文标识符", "Msg Send Type\n报文发送类型", "Msg Cycle Time (ms)\n报文周期时间", "Msg Length (Byte)\n报文长度"]
            domain = file.name.split('_')[1] + '_' + file.name.split('_')[3]
            pd_df_matrices[domain] = df

        return pd_df_matrices
    return 0

def get_dbc_2_df(dbc_files):
    if dbc_files:
        pd_df_matrices = {}
        # Получить датафреймы для каждого файла
        for file in dbc_files:
            file.seek(0)
            dbc_content = file.read().decode('utf-8')
            db = cantools.database.load_string(dbc_content, 'dbc')
            data = []
            for message in db.messages:
                data.append({
                    "Msg Name\n报文名称": message.name,
                    "Msg ID\n报文标识符": hex(message.frame_id),
                    "Msg Send Type\n报文发送类型": message.send_type,
                    "Msg Cycle Time (ms)\n报文周期时间": message.cycle_time,
                    "Msg Length (Byte)\n报文长度": message.length
                })
            df = pd.DataFrame(data)
            domain = file.name.split('_')[1] + '_' + file.name.split('_')[3]
            pd_df_matrices[domain] = df

        return pd_df_matrices
    return 0

def get_merged_df(domain_df_excel, domain_df_dbc):
    if not isinstance(domain_df_excel, int) and not isinstance(domain_df_dbc, int):
        merged_df = {**domain_df_excel, **domain_df_dbc}
    elif not isinstance(domain_df_excel, int):
        merged_df = domain_df_excel
    elif not isinstance(domain_df_dbc, int):
        merged_df = domain_df_dbc
    else:
        return 0
    return merged_df

def start_processing(merged_df):
    if merged_df:
        if st.button("Calculate busload"):
            pass
        else:
            st.stop()

def create_matrix_template(merged_df, output_path):
    if merged_df and output_path:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for domain, df in merged_df.items():
                df.to_excel(writer, sheet_name=domain, index=False)
                
        return 1
    return 0

def stylise_matrix_template(template_created, output_path):
    if template_created and output_path:
        try:
            matrix_template = load_workbook(output_path)
            grey_fill = PatternFill(start_color='D3D3D3', fill_type='solid')
            blue_fill = PatternFill(start_color='00ccff', fill_type='solid')
            for sheet_name in matrix_template.sheetnames:
                ws = matrix_template[sheet_name]
                columns = 7
                for col in range(1, columns):
                    ws.cell(row=1, column=col).fill = grey_fill
                    ws.cell(row=1, column=col).font = Font(name='Arial', size=10)
                    col_letter = get_column_letter(col)
                    first_cell = ws.cell(row=1, column=col)
                    ws.column_dimensions[col_letter].width = len(str(first_cell.value)) + 2
                    ws.cell(row=1, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column-1):
                    for cell in row:
                        cell.fill = blue_fill
            matrix_template.save(output_path)
        except Exception as e:
            st.error(f"Error occured: {str(e)}")
            st.stop()

def busload_formula_CAN500(cycle_time):
    return (134/500000)*1000/cycle_time

def busload_formula_CAN1(cycle_time):
    return (134/1000000)*1000/cycle_time

def busload_formula_CANFD2(cycle_time):
    return (30/500000+104/2000000)*1000/cycle_time

def busload_formula_CANFD5(cycle_time):
    return (30/500000+104/5000000)*1000/cycle_time

def get_domains_version(uploaded_files):
    if uploaded_files:
        domain_version = {}
        # Получить датафреймы для каждого домена
        for file in uploaded_files['xlsx']:
            df = pd.read_excel(file, sheet_name="History")
            revision_column = df.columns[0]
            domain = file.name.split('_')[1] + '_' + file.name.split('_')[3]
            domain_version[domain] = df[revision_column].dropna().iloc[-1]
        for file in uploaded_files['dbc']:
            domain = file.name.split('_')[1] + '_' + file.name.split('_')[3]
            file.seek(0)
            dbc_content = file.read().decode('utf-8')
            matches = re.findall(r'CM_ "[^"]*"', dbc_content)
            version_pattern = r'V\d\.\d\.\d'
            for match in matches:
                versions = re.findall(version_pattern, match)
                if versions:
                    last_version = versions[-1]
                else:
                    last_version = 'V1.0.0'
            domain_version[domain] = last_version

        return domain_version
    return 0

def get_recommendation(busload, recommendations):
    busload_percent = busload*100
    if busload_percent < 10:
        recommendation = recommendations['<10%']
    elif busload_percent <= 15:
        recommendation = recommendations['<=15%']
    elif busload_percent <= 30:
        recommendation = recommendations['<=30%']
    elif busload_percent <= 40:
        recommendation = recommendations['<40%']
    elif busload_percent >= 40:
        recommendation = recommendations['>=40%']
    else:
        recommendation = 'None'

    return recommendation

def get_estimation_color(busload):
    busload_percent = busload*100
    light_blue = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    light_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    green = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
    light_red = PatternFill(start_color='F4CCCC', end_color='F4CCCC', fill_type='solid')
    red = PatternFill(start_color='E26B6B', end_color='E26B6B', fill_type='solid')
    if busload_percent < 10:
        recommendation_colour = light_blue
    elif busload_percent <= 15:
        recommendation_colour = light_green
    elif busload_percent <= 30:
        recommendation_colour = green
    elif busload_percent <= 40:
        recommendation_colour = light_red
    elif busload_percent >= 40:
        recommendation_colour = red
    else:
        recommendation_colour = 'None'

    return recommendation_colour

def add_result_sheet(domain_busload, domains_version, release_version, output_path):
    if domain_busload and domains_version and output_path:
        try:
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center_alignment = Alignment(horizontal='center', vertical='center')
            recommendations = { 
                '<10%'  : "Consider decreasing speed",
                '<=15%' : "Normal (possible to decrease speed)",
                '<=30%' : "Optimal speed",
                '<40%' : "Consider increasing speed",
                '>=40%' : "BUS OFF"
            }
            speeds = ['CAN 500Kb/s', 'CAN 1Mb/s', 'CANFD 2Mb/s', 'CANFD 5Mb/s']
            header = ['Domain',	'Speed', 'Busload', 'Recommendation', 'Domain matrix version']
            df = pd.DataFrame(columns=header)
            cur_row = 1
            for domain in domain_busload:
                df.loc[cur_row] = {'Domain': domain, 'Domain matrix version': domains_version[domain]}
                for i in range(len(domain_busload[domain])):
                    speed = speeds[i]
                    busload = domain_busload[domain][i]
                    recommendation = get_recommendation(busload, recommendations)
                    if speed == 'CAN 500Kb/s':
                        if 'CANFD' in domain:
                            df.loc[cur_row+i+1] = {'Speed' : speed, 'Busload' : busload, 'Recommendation' : recommendation}
                        else:
                            df.loc[cur_row+i+1] = {'Speed' : speed, 'Busload' : busload, 'Recommendation' : recommendation, 'Domain matrix version' : 'Current speed'}
                    elif speed == 'CANFD 2Mb/s':
                        if 'CANFD' in domain:
                            df.loc[cur_row+i+1] = {'Speed' : speed, 'Busload' : busload, 'Recommendation' : recommendation, 'Domain matrix version' : 'Current speed'}
                        else:
                            df.loc[cur_row+i+1] = {'Speed' : speed, 'Busload' : busload, 'Recommendation' : recommendation}
                    else:
                        df.loc[cur_row+i+1] = {'Speed' : speed, 'Busload' : busload, 'Recommendation' : recommendation}
                    
                        
                cur_row += len(domain_busload[domain]) + 1
            with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name='Busload', index=False)
            # Переставить общий результат расчета на лицевую страницу
            busload_calculation = load_workbook(output_path)
            sheets = busload_calculation.sheetnames
            if 'Busload' in sheets:
                busload_calculation._sheets.insert(0, busload_calculation['Busload'])
                busload_calculation._sheets = [busload_calculation['Busload']] + [ws for ws in busload_calculation.worksheets if ws.title != 'Busload']
            ws = busload_calculation['Busload']
            # Покрасить строки с именем домена в серый
            for row in range(2, ws.max_row + 1, 5):  # Start at 5, step by 5
                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].fill = ws[f'A{row}'].fill = PatternFill(start_color='D3D3D3', fill_type='solid')
            # Форматировать цветом строки с рассчетом загргузки
            for row in range(2, ws.max_row + 1):
                busload_cell = ws[f'C{row}']
                if busload_cell.value is not None:
                    # Покрасить значение загрузки домена
                    busload_cell.fill = get_estimation_color(busload_cell.value)
                    busload_cell.number_format = '0.00%'
                    # Покрасить левую соседнюю ячейку со значением скорости
                    ws[f'B{row}'].fill = get_estimation_color(busload_cell.value)
                    # Покрасить правую соседнюю ячейку с рекомендацией
                    ws[f'D{row}'].fill = get_estimation_color(busload_cell.value)
            # Сгруппировать строки для возможности свертки
            for start in range(3, ws.max_row + 1, 5):  # step by 5 to leave a gap of 1 row between groups
                end = min(start + 3, ws.max_row)
                ws.row_dimensions.group(start, end, outline_level=1, hidden=False)
            # Установить ширину столбцов
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws.column_dimensions[col].width = 40
            # Применить границы и центрирование текста к ячейкам
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = border
                    cell.alignment = center_alignment
            # Добавить условные обозначения
            legend = [
                "<10% - Consider decreasing speed",
                "<=15% - Normal (possible to decrease speed)",
                "<=30% - Optimal speed",
                "<40% - Consider increasing speed",
                ">=40% - BUS OFF"
            ]
            recommendation_colours = [
                PatternFill(start_color='DDEBF7', fill_type='solid'),
                PatternFill(start_color='C6EFCE', fill_type='solid'),
                PatternFill(start_color='A9D08E', fill_type='solid'),
                PatternFill(start_color='F4CCCC', fill_type='solid'),
                PatternFill(start_color='E26B6B', fill_type='solid')
            ]
            ws.merge_cells('G1:K1')
            ws['G1'] = "Recommendations"
            ws['G1'].alignment = center_alignment
            for i, hint in enumerate(legend, start=2):
                ws.merge_cells(f'G{i}:K{i}')
                ws[f'G{i}'] = hint
                ws[f'G{i}'].fill = recommendation_colours[i-2]
                ws[f'G{i}'].alignment = center_alignment

            busload_calculation.save(output_path)

            return 1

        except Exception as e:
            st.error(f"Error occured: {str(e)}")
            st.stop()
    
    return 0

def calculate_busload(template_created, output_path):
    if template_created and output_path:
        try:
            busload_calculation = load_workbook(output_path)
            domain_busload = {}
            for sheet_name in busload_calculation.sheetnames:
                ws = busload_calculation[sheet_name]
                busload_formula = [busload_formula_CAN500, busload_formula_CAN1, busload_formula_CANFD2, busload_formula_CANFD5]
                busload_columns = ['F', 'G', 'H', 'I']
                busload_headers = [
                    'Busload CAN 500Kb/s',
                    'Busload CAN 1Mb/s',
                    'Busload CANFD 2Mb/s',
                    'Busload CANFD 5Mb/s'
                ]
                busload_sum = [0] * len(busload_columns)
                domain_busload_speeds = []
                # Установить названия колонок рассчета и ширину
                for col, header in zip(busload_columns, busload_headers):
                    ws[f'{col}1'] = header
                    ws[f'{col}1'].fill = PatternFill(start_color='D3D3D3', fill_type='solid')
                    ws[f'{col}1'].font = Font(name='Arial', size=10)
                    ws[f'{col}1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    ws.column_dimensions[col].width = 40
                # Рассчет загрузки от сообщения 
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
                    for cell in row:
                        cur_cell = cell
                        if cell.value not in (None, ""):
                            for i in range(len(busload_columns)):
                                busload_col = cell.column + 2 + i
                                busload_cell = ws.cell(row=cell.row, column=busload_col)
                                busload_cell.value = busload_formula[i](cell.value)
                                busload_cell.number_format = '0.00%'
                                busload_sum[i] += busload_cell.value
                # Заполнение ячейки загрузки от всех сообщений
                for i in range(len(busload_columns)):
                    busload_col = cur_cell.column + 2 + i
                    busload_cell = ws.cell(row=cur_cell.row+1, column=busload_col)
                    busload_cell.value = busload_sum[i]
                    busload_cell.number_format = '0.00%'
                    # Форматирование цветом ячейки с суммарной загрузкой
                    busload_cell.fill = get_estimation_color(busload_cell.value)
                    ws.cell(row=busload_cell.row+1, column=busload_col).value = busload_headers[i]
                    domain_busload_speeds.append(busload_cell.value)
                domain_busload[sheet_name] = domain_busload_speeds
            busload_calculation.save(output_path)
            busload_calculation.close()

            return domain_busload

        except Exception as e:
            st.error(f"Error occured: {str(e)}")
            st.stop()
    
    return 0

def download_busload_calculation(busload_calculated, template_path, release_version):
    if busload_calculated and template_path and release_version:
        busload_calculation = load_workbook(template_path)
        output_path = f"Busload analysis automative_ATOM_{release_version}-{datetime.now().strftime("%Y%m%d")}.xlsx"
        buffer = BytesIO()
        busload_calculation.save(buffer)
        buffer.seek(0)
        st.download_button(
            label="Download busload calculation",
            data=buffer,
            file_name=output_path,
            mime="application/octet-stream",
            type="primary",
            help="Press to download .xlsx busload calculation"
        )

def main():
    try:
        output_path = "busload_calculation_result.xlsx"
        # Озаглавить страницу
        set_page_config()
        # Предоставить форму для загрузки файлов
        uploaded_files = files_upload()
        # Предоставить поле для ввода версии релиза
        release_version = input_version()
        # Получить версии доменов
        domains_version = get_domains_version(uploaded_files)
        # Получить файлы, разделенные на xlsx и dbc форматы
        excel_files, dbc_files = get_format_splitted_files(uploaded_files)
        # Перевести загруженные таблицы в датафрейм
        domain_df_excel = get_excel_2_df(excel_files)
        # Перевести загруженные DBC в датафрейм
        domain_df_dbc = get_dbc_2_df(dbc_files)
        # Совместить excel и dbc датафреймы, убрав дублирующиеся домены
        merged_df = get_merged_df(domain_df_excel, domain_df_dbc)
        # Добавить кнопку "рассчитать"
        start_processing(merged_df)
        # Создать шаблон для рассчета загрузки
        template_created = create_matrix_template(merged_df, output_path)
        # Залить ячейки цветом
        stylise_matrix_template(template_created, output_path)
        # Рассчитать загрузку
        domain_busload = calculate_busload(template_created, output_path)
        # Добавить страницу с общим результатом
        busload_calculated = add_result_sheet(domain_busload, domains_version, release_version, output_path)
        # Скачать результат
        download_busload_calculation(busload_calculated, output_path, release_version)
        


    except Exception as e:
        st.error(f"Error occured: {str(e)}")
        st.stop()

if __name__ == "__main__":
    main()