import cantools
import cantools.database
import pandas as pd
import traceback
import argparse
import pprint
import os
from openpyxl import load_workbook
from typing import List, Dict
from collections import OrderedDict


class DbcRead:
    def __init__(self, dbc_path: str):
        self.dbc_path = dbc_path

    def CreateDB(self):
        db = cantools.database.load_file(self.dbc_path)

        result = {}

        for message in db.messages:
            message_attr = message.dbc.attributes
            gen_msg_cycle_time_fast = (
                message_attr.get("GenMsgCycleTimeFast").value
                if message_attr.get("GenMsgCycleTimeFast")
                else None
            )
            gen_msg_nr_rep = (
                message_attr.get("GenMsgNrOfRepetition").value
                if message_attr.get("GenMsgNrOfRepetition")
                else None
            )
            gen_msg_delay = (
                message_attr.get("GenMsgDelayTime").value
                if message_attr.get("GenMsgDelayTime")
                else None
            )
            result[message.name] = {
                "Msg_id": message.frame_id,
                "Msg_length": message.length,
                "Bus_name": message.bus_name,
                "Msg_comment": message.comment,
                "Cycle_time": message.cycle_time,
                "Byte_order": message.header_byte_order,
                "If_fd": message.is_fd,
                "Protocol": message.protocol,
                "Send_type": message.send_type,
                "Senders": message.senders,
                "Recievers": message.receivers,
                "GenMsgCycleTimeFast": gen_msg_cycle_time_fast,
                "GenMsgNrOfRepetition": gen_msg_nr_rep,
                "GenMsgDelayTime": gen_msg_delay,
                "Signals": [],
            }

            for signal in message.signals:
                signal_data = {
                    "Sgn_name": signal.name,
                    "Start_bit": signal.start,
                    "Sgn_lenght": signal.length,
                    "Byte_oreder": signal.byte_order,
                    "Sgn_Send_Type": signal.dbc.attributes["GenSigSendType"],
                    "Is_signed": signal.is_signed,
                    "Initinal": signal.raw_initial,
                    "Invalid": signal.raw_invalid,
                    "Factor": signal.conversion.scale,
                    "Offset": signal.conversion.offset,
                    "Is_float": signal.conversion.is_float,
                    "Minimum": signal.minimum,
                    "Maximum": signal.maximum,
                    "Unit": signal.unit,
                    "Comment": signal.comment,
                    "Receivers": signal.receivers,
                    "Value_description": signal.choices,
                }
                result[message.name]["Signals"].append(signal_data)
        bus_users = db.nodes

        return result, bus_users

    def _format_value_description(self, choices):
        if not choices or choices == None:
            return ""

        if isinstance(choices, (dict, OrderedDict)):
            lines = []
            for value, desc in choices.items():
                hex_value = f"0x{int(value):X}"
                lines.append(f"{hex_value}: {desc}")
            return "\n".join(lines)
        return str(choices)

    def copy_format(self, source_path: str, target_path: str) -> None:
        """Copy styles from source.xlsx in target.xlsx."""
        try:
            source_wb = load_workbook(source_path)
            target_wb = load_workbook(target_path)

            source_sheet_name = source_wb.sheetnames[0]
            target_sheet_name = target_wb.sheetnames[0]

            source_sheet = source_wb[source_sheet_name]
            target_sheet = target_wb[target_sheet_name]

            for row in source_sheet.iter_rows():
                for cell in row:
                    target_cell = target_sheet[cell.coordinate]

                    target_cell.number_format = cell.number_format

            target_wb.save(target_path)
            print(f"Formating successfully copy from {source_path} in {target_path}")
        except Exception as e:
            print(f"Error durring copy formats: {str(e)}")

    def convert(self, output_path: str = "output.xlsx") -> bool:
        """Main method convert (message row + signal rows, with style copy)"""
        try:
            print(f"Starting conversion to: {output_path}")
            print(f"Current working directory: {os.getcwd()}")
            print(f"DBC file path: {self.dbc_path}")
            print(f"DBC file exists: {os.path.exists(self.dbc_path)}")
            lib, ecu = self.CreateDB()
            ecu_nodes = [node.name for node in ecu]

            test_xlsx_path = None
            possible_paths = [
                "test.xlsx",
                "pages/test.xlsx",
                os.path.join(os.path.dirname(__file__), "test.xlsx"),
                os.path.join(os.path.dirname(__file__), "pages", "test.xlsx"),
            ]

            for path in possible_paths:
                if os.path.exists(path):
                    test_xlsx_path = path
                    break

            if not test_xlsx_path:
                print("Warning: test.xlsx not found, using default columns")
                base_columns = [
                    "Message Name",
                    "Message Type",
                    "Message ID",
                    "Send Type",
                    "Cycle Time",
                    "Protocol",
                    "CAN FD",
                    "Message Length",
                    "Signal Name",
                    "Signal Description",
                    "Byte Order",
                    "Start Byte",
                    "Start Bit",
                    "Send Type",
                    "Bit Length",
                    "Data Type",
                    "Resolution",
                    "Offset",
                    "Min Value",
                    "Max Value",
                    "Min Raw",
                    "Max Raw",
                    "Initial Value",
                    "Invalid Value",
                    "Error Value",
                    "Unit",
                    "Value Description",
                    "GenMsgCycleTimeFast",
                    "GenMsgNrOfRepetition",
                    "GenMsgDelayTime",
                ]
                test_sheet_name = "Sheet1"
            else:
                test_xl = pd.ExcelFile(test_xlsx_path)
                test_sheet_name = test_xl.sheet_names[0]
                test_df = pd.read_excel(test_xl, sheet_name=test_sheet_name)
                base_columns = list(map(str, test_df.columns))

            non_ecu_columns = []
            for col in base_columns:
                if col not in ecu_nodes:
                    non_ecu_columns.append(col)
            columns = non_ecu_columns + ecu_nodes

            rows = []
            for msg_name, msg_data in lib.items():
                msg_row = {col: "" for col in columns}
                msg_row[columns[0]] = msg_name
                msg_row[columns[1]] = (
                    "NM"
                    if str(msg_name).startswith("NM_")
                    else "Diag" if str(msg_name).startswith("Diag") else "Normal"
                )
                msg_row[columns[2]] = f"0x{int(msg_data['Msg_id']):X}"
                msg_row[columns[3]] = msg_data.get("Send_type", "")
                msg_row[columns[4]] = msg_data["Cycle_time"]
                msg_row[columns[5]] = (
                    "StandardCAN" if msg_data["Protocol"] == "CAN" else "StandardCAN_FD"
                )
                msg_row[columns[6]] = (
                    str(1) if msg_row[columns[5]] == "StandardCAN_FD" else str(0)
                )
                msg_row[columns[7]] = msg_data["Msg_length"]

                for ecu_node in ecu_nodes:
                    if ecu_node in columns:
                        msg_row[ecu_node] = (
                            "S" if ecu_node in msg_data["Senders"] else "R"
                        )
                rows.append(msg_row)

                for signal in msg_data["Signals"]:
                    sig_row = {col: "" for col in columns}
                    sig_row[columns[8]] = signal["Sgn_name"]
                    sig_row[columns[9]] = signal["Comment"]
                    sig_row[columns[10]] = (
                        "Motorola MSB"
                        if signal["Byte_oreder"] == "big_endian"
                        else "Intel"
                    )
                    sig_row[columns[11]] = signal["Start_bit"] // 8
                    sig_row[columns[12]] = signal["Start_bit"]
                    send_map = {
                        0: "Cyclic",
                        1: "OnChange",
                        2: "OnWrite",
                        3: "IfActive",
                        4: "OnChangeWithRepetition",
                        5: "OnWriteWithRepetition",
                        6: "IfActiveWithRepetition",
                        7: "NoSigSendType",
                        8: "OnChangeAndIfActive",
                        9: "OnChangeAndIfActiveWithRepetition",
                        10: "CA",
                        11: "CE",
                        12: "Event",
                    }
                    gen_sig_send_type = (
                        signal["Sgn_Send_Type"].value
                        if hasattr(signal["Sgn_Send_Type"], "value")
                        else None
                    )
                    sig_row[columns[13]] = send_map.get(gen_sig_send_type, "")
                    sig_row[columns[14]] = signal["Sgn_lenght"]
                    sig_row[columns[15]] = (
                        "Unsigned" if signal["Is_signed"] == False else "Signed"
                    )
                    sig_row[columns[16]] = signal["Factor"]
                    sig_row[columns[17]] = signal["Offset"]
                    sig_row[columns[18]] = signal["Minimum"]
                    sig_row[columns[19]] = signal["Maximum"]
                    sig_row[columns[20]] = (
                        f"0x{int((signal['Minimum'] - signal['Offset']) / signal['Factor']):X}"
                        if signal["Factor"] != 0
                        else "0x0"
                    )
                    sig_row[columns[21]] = (
                        f"0x{int((signal['Maximum'] - signal['Offset']) / signal['Factor']):X}"
                        if signal["Factor"] != 0
                        else "0x0"
                    )
                    sig_row[columns[22]] = (
                        f"0x{int(signal['Initinal']):X}"
                        if pd.notna(signal["Initinal"])
                        else ""
                    )
                    sig_row[columns[23]] = (
                        f"0x{int(signal['Invalid']):X}"
                        if pd.notna(signal["Invalid"])
                        else ""
                    )
                    sig_row[columns[24]] = "0x0"
                    sig_row[columns[25]] = signal["Unit"]

                    sig_row[columns[26]] = self._format_value_description(
                        signal["Value_description"]
                    )
                    sig_row[columns[27]] = msg_data["GenMsgCycleTimeFast"]
                    sig_row[columns[28]] = msg_data["GenMsgNrOfRepetition"]
                    sig_row[columns[29]] = msg_data["GenMsgDelayTime"]

                    for ecu_node in ecu_nodes:
                        if ecu_node in columns:
                            if ecu_node in msg_data["Senders"]:
                                sig_row[ecu_node] = "S"
                            elif ecu_node in signal["Receivers"]:
                                sig_row[ecu_node] = "R"
                            else:
                                sig_row[ecu_node] = ""
                    rows.append(sig_row)

            df = pd.DataFrame(rows, columns=columns)
            df.to_excel(
                output_path, sheet_name=test_sheet_name, index=False, engine="openpyxl"
            )

            if test_xlsx_path:
                self.copy_format(test_xlsx_path, output_path)

            print(f"Excel file successfully created: {output_path}")
            return True
        except Exception as e:
            print(f"Error during conversion: {str(e)}")
            print(f"Error type: {type(e)}")
            print(f"Full traceback: {traceback.format_exc()}")
            return False


if __name__ == "__main__":
    converter = DbcRead(
        "C:\\projects\\Convert2DBC\\ATOM_CAN_Matrix_BD_V8.0.0_20250625.dbc"
    )
    converter.convert("output.xlsx")
