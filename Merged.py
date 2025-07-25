import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from copy import copy
import os
from pathlib import Path
from io import BytesIO
from datetime import datetime
# Формирование папок
import create_directory
# Замер времени выполнения блоков
import time
import concurrent.futures

def set_page_title():
    st.title("Release Convertor")

def get_uploaded_file():
    uploaded_file = st.file_uploader("Upload Domain Excel Matrix", type=["xlsx"])
    if uploaded_file:
        return uploaded_file

def process_matrix_sheet(wb, ecu_col_index):

    process_matrix_sheet_time_start = time.time()

    new_matrix_ws = wb["Matrix"]
    # Убрать группировки для строк и столбцов для корректного форматирования при удалении строк
    for row in range(1, new_matrix_ws.max_row + 1):
        new_matrix_ws.row_dimensions[row].outline_level = 0
        new_matrix_ws.row_dimensions[row].hidden = False
    # Удалить строки, не соответствующие выбанному ecu
    for row_idx in range(new_matrix_ws.max_row, 1, -1):  # from max_row down to 2
        row = list(new_matrix_ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
        if not any((row[idx].value is not None and str(row[idx].value).lower() in ["s", "r"]) for idx in ecu_col_index.values()):
            new_matrix_ws.delete_rows(row_idx, amount=1)
    # Группировать строки по сообщениям
    row_idx = 2
    while row_idx <= new_matrix_ws.max_row:
        a_val = new_matrix_ws.cell(row=row_idx, column=1).value
        new_matrix_ws.row_dimensions[row_idx].height = 20
        if a_val:
            start_row = row_idx + 1
            end_row = start_row
            while end_row <= new_matrix_ws.max_row:
                i_val = new_matrix_ws.cell(row=end_row, column=9).value
                if i_val:
                    end_row += 1
                else:
                    break
            if end_row > start_row:
                for r in range(start_row, end_row):
                    new_matrix_ws.row_dimensions[r].outlineLevel = 1
                    new_matrix_ws.row_dimensions[r].hidden = True
                new_matrix_ws.row_dimensions[row_idx].collapsed = True
            row_idx = end_row
        else:
            row_idx += 1
    # Удалить лишние данные
    max_rows, max_cols = 5000, 50
    if new_matrix_ws.max_row > max_rows:
        new_matrix_ws.delete_rows(max_rows + 1, new_matrix_ws.max_row - max_rows)
    if new_matrix_ws.max_column > max_cols:
        new_matrix_ws.delete_cols(max_cols + 1, new_matrix_ws.max_column - max_cols)

    process_matrix_sheet_time_end = time.time() - process_matrix_sheet_time_start
    print("process_matrix_sheet_time_end", process_matrix_sheet_time_end)

    return wb

def process_history_sheet(wb, ecu_col_index):

    process_history_sheet_time_start = time.time()

    if "History" not in wb.sheetnames:

        return False
    # Удалить строки, не соответствующие выбанному ecu
    ecu_history_ws = wb['History']
    history_col_idx = 6
    ecu_name = next(iter(ecu_col_index))
    for row_idx in range(ecu_history_ws.max_row, 1, -1):
        cell_value = ecu_history_ws.cell(row=row_idx, column=history_col_idx).value
        if cell_value is None or ecu_name not in str(cell_value):
            ecu_history_ws.delete_rows(row_idx)
    
    process_history_sheet_time_end = time.time() - process_history_sheet_time_start
    print("process_history_sheet_time_end", process_history_sheet_time_end)

    return ecu_wb

def identify_ecus(df):
    # Найти все ecu в предоставленной доменной матрице
    ecus = [
        col
        for col in df.columns
        if any(val in ["S", "R"] for val in df[col].dropna().unique())
        and col != "Unit\n单位"
    ]

    if not ecus:
        st.error(
            "🕭 No ECUs found in the file. Please check that the file contains columns with 'S' or 'R' values."
        )
        st.stop()

    return ecus

def get_ecu_version(wb, ecu_col_index):

    get_ecu_version_time_start = time.time()
    
    ecu_history_ws = wb['History']
    history_col_idx = 6
    ecu_name = next(iter(ecu_col_index))
    # Найти версию по первому снизу вхождению выбранного ecu
    for row_idx in range(ecu_history_ws.max_row, 1, -1):
        cell_value = ecu_history_ws.cell(row=row_idx, column=history_col_idx).value
        if ecu_name in str(cell_value):
            ecu_version = ecu_history_ws.cell(row=row_idx, column=1).value
            break
    
    get_ecu_version_time_end = time.time() - get_ecu_version_time_start
    print("get_ecu_version_time_end", get_ecu_version_time_end)

    return ecu_version

def get_domain_folder_name(ecu_base, domain_short):
    if ecu_base in ("SGW", "CGW", "ADCU"):
        domain_short = "SGW-CGW"
    domain_folder_name = {
        "BD" :      "04.01.01.Body CAN",
        "DG" :      "",
        "CH" :      "04.01.05.Chassis CANFD",
        "PT" :      "04.01.02.Powertrain CAN",
        "ET" :      "04.01.04.Entertainment CANFD",
        "DZ" :      "04.01.06.Demilitary zone CANFD",
        "SGW-CGW" : "04.01.07.CGW,SGW,ADCU"
    }

    return domain_folder_name[domain_short]

def get_ecu_folder_name(domain_folder_name, ecu_base):
    ecu_folder_name = ""
    for sub_folder_name in create_directory.creator.HIERARCHI[domain_folder_name]:
        print(ecu_base, sub_folder_name)
        if ecu_base in sub_folder_name:
            ecu_folder_name = sub_folder_name
            break
    if ecu_folder_name:
        return ecu_folder_name
    else:
        print("Sub folder for ECU doesn't match ECU name.", )
        # st.error("Sub folder for ECU doesn't match ECU name.")

if __name__ == "__main__":
    set_page_title()
    uploaded_file = get_uploaded_file()
    print("File uploaded")
    # Доменный excel получен, дальше блок обработки для получения данных и распределение по папкам 
    if uploaded_file:
        try:
            domain_short = uploaded_file.name.split('_')[3]
            df = pd.read_excel(uploaded_file, sheet_name="Matrix")
            ecus = identify_ecus(df)
            for ecu in ecus:

                total_start = time.time()

                date_str = datetime.now().strftime("%d%m%Y")
                ecu_col_index = {ecu: df.columns.get_loc(ecu)}
                # Получить основу имени ecu
                if '_' in ecu:
                    ecu_base = ecu.split("_")[0]
                else:
                    ecu_base = ecu
                # Создать копию доменной матрицы для ее модификации в матрицу ecu
                ecu_wb = load_workbook(uploaded_file)
                # Настроить лист "Matrix" под ecu
                ecu_wb = process_matrix_sheet(ecu_wb, ecu_col_index)
                ecu_wb = process_history_sheet(ecu_wb, ecu_col_index)
                if not ecu_wb:
                    st.warning(
                        f"🕱 'History' sheet not found in ECU {ecu}."
                    )
                ecu_version = get_ecu_version(ecu_wb, ecu_col_index)
                # Сохранить ecu матрицу в папку

                find_path_time_start = time.time()

                domain_folder_name = get_domain_folder_name(ecu_base, domain_short)
                ecu_folder_name = get_ecu_folder_name(domain_folder_name, ecu_base)
                output_ecu_filename = f"ATOM_CAN_MATRIX_{ecu_version}_{date_str}_{ecu}.xlsx"
                ecu_matrix_output_path = f"{create_directory.creator.PATH_DOC}\\{domain_folder_name}\\{ecu_folder_name}\\{output_ecu_filename}"
                
                find_path_time_end = time.time() - find_path_time_start
                print("find_path_time_end", find_path_time_end)

                ecu_save_time_start = time.time()

                ecu_wb.save(ecu_matrix_output_path)

                ecu_save_time_end = time.time() - ecu_save_time_start
                print("ecu_save_time_end", ecu_save_time_end)

                ecu_proccesed_time = time.time() - total_start
                print("ecu_proccesed_time", ecu_proccesed_time)

            st.success(
                f"Domain matrix ecu split completed, obtained {len(ecus)} ECUs."
            )
        except Exception as e:
            st.error(f"🕭 Error processing file: {e}")
